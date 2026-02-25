#!/usr/bin/env python3
"""Tests for swap_operands_perf_compare.py.

Run with: pytest test_swap_operands_perf_compare.py -v
"""

import math
import os
import tempfile
from pathlib import Path
from unittest import mock

import numpy as np
import pandas as pd
import pytest

from swap_operands_perf_compare import (
    ComparisonRow,
    ExcelMetadata,
    PerfEntry,
    TuningEntry,
    _parse_conv_flags,
    _parse_cpu_list,
    compute_diff_pct,
    conv_row_to_test_vector,
    detect_gpus_rocm_smi,
    detect_num_cpus,
    extract_global_store_width_from_isa,
    gemm_row_to_test_vector,
    generate_excel,
    merge_results,
    parse_perf_csv,
    parse_tuning_tsv,
    sort_comparison_rows,
    validate_excel,
    validate_excel_against_sources,
)

# ---------------------------------------------------------------------------
# TSV Parsing Tests
# ---------------------------------------------------------------------------


class TestParseTuningTsv:

    def _write_tsv(self, tmp_path: Path, content: str) -> str:
        p = tmp_path / "tuning.tsv"
        p.write_text(content)
        return str(p)

    def test_v3_format(self, tmp_path):
        content = (
            "# commit: abc123\n"
            "# arch\tnumCUs\tnumChiplets\ttestVector\tperfConfig (greedy)\n"
            "gfx950:sramecc+:xnack-\t256\t8\t"
            "-t f32 -out_datatype f32 -transA false -transB false -g 1 -m 384 -n 768 -k 3072\t"
            "v4:32,32,8,16,16,16,4,3,4,0,8,0,1,1\n")
        path = self._write_tsv(tmp_path, content)
        result = parse_tuning_tsv(path)
        assert len(result) == 1
        tv = "-t f32 -out_datatype f32 -transA false -transB false -g 1 -m 384 -n 768 -k 3072"
        assert tv in result
        entry = result[tv]
        assert entry.arch == "gfx950:sramecc+:xnack-"
        assert entry.num_cu == 256
        assert entry.num_chiplets == 8
        assert entry.perf_config == "v4:32,32,8,16,16,16,4,3,4,0,8,0,1,1"

    def test_v2_format(self, tmp_path):
        content = (
            "# header\n"
            "gfx90a\t120\t-t f16 -out_datatype f16 -transA false -transB false -g 1 -m 100 -n 200 -k 300\t"
            "v4:16,16,4,16,16,16,4,1,4,0,1,32,1,1\n")
        path = self._write_tsv(tmp_path, content)
        result = parse_tuning_tsv(path)
        assert len(result) == 1
        tv = "-t f16 -out_datatype f16 -transA false -transB false -g 1 -m 100 -n 200 -k 300"
        assert tv in result
        assert result[tv].num_cu == 120
        assert result[tv].num_chiplets == 0

    def test_legacy_format(self, tmp_path):
        content = (
            "gfx908\t-t f32 -out_datatype f32 -transA false -transB false -g 1 -m 10 -n 20 -k 30\t"
            "v4:8,8,4,8,8,8,4,1,4,0,1,0,1,1\n")
        path = self._write_tsv(tmp_path, content)
        result = parse_tuning_tsv(path)
        assert len(result) == 1

    def test_skips_comments_and_empty_lines(self, tmp_path):
        content = ("# comment line 1\n"
                   "\n"
                   "# comment line 2\n"
                   "gfx950:sramecc+:xnack-\t256\t8\t"
                   "-t f32 -out_datatype f32 -transA false -transB false -g 1 -m 1 -n 2 -k 3\t"
                   "v4:1,1,1,1,1,1,1,1,1,0,1,0,1,1\n"
                   "\n")
        path = self._write_tsv(tmp_path, content)
        result = parse_tuning_tsv(path)
        assert len(result) == 1

    def test_multiple_entries(self, tmp_path):
        content = (
            "# header\n"
            "gfx950\t256\t8\t-t f32 -out_datatype f32 -transA false -transB false -g 1 -m 10 -n 20 -k 30\tv4:a\n"
            "gfx950\t256\t8\t-t f16 -out_datatype f16 -transA false -transB true -g 1 -m 40 -n 50 -k 60\tv4:b\n"
            "gfx950\t256\t8\t-t i8 -out_datatype i32 -transA true -transB false -g 2 -m 70 -n 80 -k 90\tv4:c\n"
        )
        path = self._write_tsv(tmp_path, content)
        result = parse_tuning_tsv(path)
        assert len(result) == 3


# ---------------------------------------------------------------------------
# CSV Parsing Tests
# ---------------------------------------------------------------------------


class TestParsePerfCsv:

    def _write_gemm_csv(self, tmp_path: Path, rows: list) -> str:
        p = tmp_path / "perf.csv"
        headers = [
            'DataType', 'OutDataType', 'Chip', 'numCU', 'numChiplets', 'TransA', 'TransB', 'G', 'M',
            'K', 'N', 'ScaledGemm', 'ScaleADtype', 'ScaleBDtype', 'TransScaleA', 'TransScaleB',
            'PerfConfig', 'LDSBankConflict', 'TFlops'
        ]
        df = pd.DataFrame(rows, columns=headers)
        df.to_csv(str(p), index=False)
        return str(p)

    def _write_conv_csv(self, tmp_path: Path, rows: list) -> str:
        p = tmp_path / "perf.csv"
        headers = [
            'Direction', 'DataType', 'Chip', 'numCU', 'numChiplets', 'FilterLayout', 'InputLayout',
            'OutputLayout', 'N', 'C', 'H', 'W', 'K', 'Y', 'X', 'DilationH', 'DilationW', 'StrideH',
            'StrideW', 'PaddingH', 'PaddingW', 'PerfConfig', 'LDSBankConflict', 'TFlops'
        ]
        df = pd.DataFrame(rows, columns=headers)
        df.to_csv(str(p), index=False)
        return str(p)

    def test_gemm_csv_parsing(self, tmp_path):
        rows = [[
            'f32', 'f32', 'gfx950', 256, 8, 'false', 'false', 1, 384, 3072, 768, 'False', '', '',
            'False', 'False', 'v4:32,32,8,16,16,16,4,3,4,0,8,0,1,1', 0, 42.5
        ]]
        path = self._write_gemm_csv(tmp_path, rows)
        result = parse_perf_csv(path, 'gemm')
        assert len(result) == 1
        key = list(result.keys())[0]
        assert '-t f32' in key
        assert '-m 384' in key
        assert '-n 768' in key
        assert '-k 3072' in key
        assert result[key].tflops == 42.5
        assert 'v4:32,32' in result[key].perf_config

    def test_gemm_csv_test_vector_reconstruction(self, tmp_path):
        rows = [[
            'f16', 'f16', 'gfx950', 256, 8, 'true', 'false', 12, 384, 64, 384, 'False', '', '',
            'False', 'False', 'v4:x', 0, 10.0
        ]]
        path = self._write_gemm_csv(tmp_path, rows)
        result = parse_perf_csv(path, 'gemm')
        key = list(result.keys())[0]
        assert '-t f16' in key
        assert '-transA true' in key
        assert '-transB false' in key
        assert '-g 12' in key

    def test_conv_csv_parsing(self, tmp_path):
        rows = [[
            'fwd', 'f32', 'gfx950', 256, 8, 'gkcyx', 'ngchw', 'ngkhw', 1, 256, 14, 14, 1024, 1, 1,
            1, 1, 1, 1, 0, 0, 'v4:cfg', 0, 35.7
        ]]
        path = self._write_conv_csv(tmp_path, rows)
        result = parse_perf_csv(path, 'conv')
        assert len(result) == 1
        key = list(result.keys())[0]
        assert result[key].tflops == 35.7

    def test_multiple_gemm_entries(self, tmp_path):
        rows = [
            [
                'f32', 'f32', 'gfx950', 256, 8, 'false', 'false', 1, 384, 3072, 768, 'False', '',
                '', 'False', 'False', 'v4:a', 0, 42.5
            ],
            [
                'f16', 'f16', 'gfx950', 256, 8, 'false', 'true', 1, 100, 200, 300, 'False', '', '',
                'False', 'False', 'v4:b', 0, 55.0
            ],
        ]
        path = self._write_gemm_csv(tmp_path, rows)
        result = parse_perf_csv(path, 'gemm')
        assert len(result) == 2

    def test_nan_tflops(self, tmp_path):
        rows = [[
            'f32', 'f32', 'gfx950', 256, 8, 'false', 'false', 1, 10, 20, 30, 'False', '', '',
            'False', 'False', 'v4:z', 0,
            float('nan')
        ]]
        path = self._write_gemm_csv(tmp_path, rows)
        result = parse_perf_csv(path, 'gemm')
        key = list(result.keys())[0]
        assert math.isnan(result[key].tflops)


# ---------------------------------------------------------------------------
# Test Vector Reconstruction Tests
# ---------------------------------------------------------------------------


class TestGemmRowToTestVector:

    def test_basic(self):
        row = {
            'DataType': 'f32',
            'OutDataType': 'f32',
            'TransA': 'false',
            'TransB': 'false',
            'G': 1,
            'M': 384,
            'N': 768,
            'K': 3072,
            'ScaledGemm': 'False',
            'ScaleADtype': '',
            'ScaleBDtype': '',
            'TransScaleA': 'False',
            'TransScaleB': 'False',
        }
        tv = gemm_row_to_test_vector(row)
        assert tv == "-t f32 -out_datatype f32 -transA false -transB false -g 1 -m 384 -n 768 -k 3072"

    def test_with_trans(self):
        row = {
            'DataType': 'f16',
            'OutDataType': 'f16',
            'TransA': 'True',
            'TransB': 'False',
            'G': 12,
            'M': 384,
            'N': 64,
            'K': 384,
            'ScaledGemm': 'False',
            'ScaleADtype': '',
            'ScaleBDtype': '',
            'TransScaleA': 'False',
            'TransScaleB': 'False',
        }
        tv = gemm_row_to_test_vector(row)
        assert '-transA true' in tv
        assert '-transB false' in tv


# ---------------------------------------------------------------------------
# ISA Parsing Tests
# ---------------------------------------------------------------------------


class TestExtractGlobalStoreWidth:

    def test_dword(self):
        isa = "global_store_dword v[0:1], v2, off\n"
        assert extract_global_store_width_from_isa(isa) == 4

    def test_dwordx2(self):
        isa = "global_store_dwordx2 v[0:1], v[2:3], off\n"
        assert extract_global_store_width_from_isa(isa) == 8

    def test_dwordx4(self):
        isa = "global_store_dwordx4 v[0:1], v[2:5], off\n"
        assert extract_global_store_width_from_isa(isa) == 16

    def test_byte(self):
        isa = "global_store_byte v[0:1], v2, off\n"
        assert extract_global_store_width_from_isa(isa) == 1

    def test_short(self):
        isa = "global_store_short v[0:1], v2, off\n"
        assert extract_global_store_width_from_isa(isa) == 2

    def test_max_of_multiple(self):
        isa = ("global_store_dword v[0:1], v2, off\n"
               "global_store_dwordx4 v[3:4], v[5:8], off\n"
               "global_store_byte v[9:10], v11, off\n")
        assert extract_global_store_width_from_isa(isa) == 16

    def test_no_stores(self):
        isa = "s_load_dword s0, s[0:1], 0\nv_add_f32 v0, v1, v2\n"
        assert extract_global_store_width_from_isa(isa) == 0

    def test_empty(self):
        assert extract_global_store_width_from_isa("") == 0

    def test_b128_format(self):
        isa = "global_store_b128 v[0:1], v[2:5], off\n"
        assert extract_global_store_width_from_isa(isa) == 16

    def test_b32_format(self):
        isa = "global_store_b32 v[0:1], v2, off\n"
        assert extract_global_store_width_from_isa(isa) == 4

    def test_dwordx3(self):
        isa = "global_store_dwordx3 v[0:1], v[2:4], off\n"
        assert extract_global_store_width_from_isa(isa) == 12

    def test_mixed_old_and_new_format(self):
        isa = ("global_store_dword v[0:1], v2, off\n"
               "global_store_b128 v[3:4], v[5:8], off\n")
        assert extract_global_store_width_from_isa(isa) == 16


# ---------------------------------------------------------------------------
# Diff Calculation Tests
# ---------------------------------------------------------------------------


class TestComputeDiffPct:

    def test_positive_improvement(self):
        diff = compute_diff_pct(100.0, 110.0)
        assert abs(diff - 10.0) < 0.001

    def test_negative_regression(self):
        diff = compute_diff_pct(100.0, 90.0)
        assert abs(diff - (-10.0)) < 0.001

    def test_no_change(self):
        diff = compute_diff_pct(100.0, 100.0)
        assert abs(diff) < 0.001

    def test_base_zero(self):
        diff = compute_diff_pct(0.0, 10.0)
        assert math.isnan(diff)

    def test_base_nan(self):
        diff = compute_diff_pct(float('nan'), 10.0)
        assert math.isnan(diff)

    def test_feature_nan(self):
        diff = compute_diff_pct(10.0, float('nan'))
        assert math.isnan(diff)

    def test_both_nan(self):
        diff = compute_diff_pct(float('nan'), float('nan'))
        assert math.isnan(diff)

    def test_large_improvement(self):
        diff = compute_diff_pct(10.0, 30.0)
        assert abs(diff - 200.0) < 0.001

    def test_small_values(self):
        diff = compute_diff_pct(0.001, 0.002)
        assert abs(diff - 100.0) < 0.1


# ---------------------------------------------------------------------------
# Merge and Sort Tests
# ---------------------------------------------------------------------------


class TestMergeResults:

    def _make_tuning(self, tv, pc):
        return TuningEntry("gfx950", 256, 8, tv, pc)

    def _make_perf(self, tv, pc, tflops):
        return PerfEntry(tv, pc, tflops)

    def test_basic_merge(self):
        base_t = {"cfg1": self._make_tuning("cfg1", "pc_b1")}
        feat_t = {"cfg1": self._make_tuning("cfg1", "pc_f1")}
        base_p = {"cfg1": self._make_perf("cfg1", "pc_b1", 100.0)}
        feat_p = {"cfg1": self._make_perf("cfg1", "pc_f1", 110.0)}
        base_i = {"cfg1": 4}
        feat_i = {"cfg1": 16}

        rows = merge_results(base_t, feat_t, base_p, feat_p, base_i, feat_i)
        assert len(rows) == 1
        assert rows[0].tflops_base == 100.0
        assert rows[0].tflops_feature == 110.0
        assert rows[0].store_width_base == 4
        assert rows[0].store_width_feature == 16
        assert abs(rows[0].diff_pct - 10.0) < 0.001

    def test_missing_in_feature(self):
        base_t = {"cfg1": self._make_tuning("cfg1", "pc1")}
        feat_t = {}
        base_p = {"cfg1": self._make_perf("cfg1", "pc1", 100.0)}
        feat_p = {}

        rows = merge_results(base_t, feat_t, base_p, feat_p, {}, {})
        assert len(rows) == 1
        assert rows[0].perf_config_feature == ""
        assert math.isnan(rows[0].tflops_feature)

    def test_missing_in_base(self):
        base_t = {}
        feat_t = {"cfg1": self._make_tuning("cfg1", "pc1")}
        base_p = {}
        feat_p = {"cfg1": self._make_perf("cfg1", "pc1", 100.0)}

        rows = merge_results(base_t, feat_t, base_p, feat_p, {}, {})
        assert len(rows) == 1
        assert rows[0].perf_config_base == ""
        assert math.isnan(rows[0].tflops_base)

    def test_union_of_keys(self):
        base_t = {"cfg1": self._make_tuning("cfg1", "a"), "cfg2": self._make_tuning("cfg2", "b")}
        feat_t = {"cfg2": self._make_tuning("cfg2", "c"), "cfg3": self._make_tuning("cfg3", "d")}
        base_p = {
            "cfg1": self._make_perf("cfg1", "a", 10.0),
            "cfg2": self._make_perf("cfg2", "b", 20.0)
        }
        feat_p = {
            "cfg2": self._make_perf("cfg2", "c", 25.0),
            "cfg3": self._make_perf("cfg3", "d", 30.0)
        }

        rows = merge_results(base_t, feat_t, base_p, feat_p, {}, {})
        configs = {r.problem_config for r in rows}
        assert configs == {"cfg1", "cfg2", "cfg3"}


class TestSortComparisonRows:

    def _row(self, config, diff):
        return ComparisonRow(config, "", "", 0, 0, 0, 0, diff)

    def test_positive_before_negative(self):
        rows = [
            self._row("a", -5.0),
            self._row("b", 10.0),
            self._row("c", 3.0),
            self._row("d", -2.0),
        ]
        sorted_rows = sort_comparison_rows(rows)
        assert sorted_rows[0].diff_pct == 10.0
        assert sorted_rows[1].diff_pct == 3.0
        assert sorted_rows[2].diff_pct == -2.0
        assert sorted_rows[3].diff_pct == -5.0

    def test_nan_at_end(self):
        rows = [
            self._row("a", float('nan')),
            self._row("b", 5.0),
            self._row("c", -3.0),
        ]
        sorted_rows = sort_comparison_rows(rows)
        assert sorted_rows[0].diff_pct == 5.0
        assert sorted_rows[1].diff_pct == -3.0
        assert math.isnan(sorted_rows[2].diff_pct)

    def test_all_positive_descending(self):
        rows = [
            self._row("a", 1.0),
            self._row("b", 5.0),
            self._row("c", 3.0),
        ]
        sorted_rows = sort_comparison_rows(rows)
        assert [r.diff_pct for r in sorted_rows] == [5.0, 3.0, 1.0]

    def test_all_negative_descending(self):
        rows = [
            self._row("a", -1.0),
            self._row("b", -5.0),
            self._row("c", -3.0),
        ]
        sorted_rows = sort_comparison_rows(rows)
        assert [r.diff_pct for r in sorted_rows] == [-1.0, -3.0, -5.0]

    def test_empty(self):
        assert sort_comparison_rows([]) == []

    def test_single(self):
        rows = [self._row("a", 42.0)]
        sorted_rows = sort_comparison_rows(rows)
        assert len(sorted_rows) == 1
        assert sorted_rows[0].diff_pct == 42.0

    def test_zero_diff(self):
        rows = [
            self._row("a", 0.0),
            self._row("b", 5.0),
            self._row("c", -3.0),
        ]
        sorted_rows = sort_comparison_rows(rows)
        assert sorted_rows[0].diff_pct == 5.0
        assert sorted_rows[1].diff_pct == 0.0
        assert sorted_rows[2].diff_pct == -3.0


# ---------------------------------------------------------------------------
# Excel Generation and Validation Tests
# ---------------------------------------------------------------------------


class TestExcelGeneration:

    def _make_rows(self):
        return [
            ComparisonRow("cfg_a", "pc_base_a", "pc_feat_a", 100.0, 110.0, 4, 16, 10.0),
            ComparisonRow("cfg_b", "pc_base_b", "pc_feat_b", 50.0, 45.0, 8, 8, -10.0),
            ComparisonRow("cfg_c", "pc_base_c", "pc_feat_c", 80.0, 80.0, 16, 16, 0.0),
        ]

    @pytest.mark.skipif(not pytest.importorskip("openpyxl"), reason="openpyxl required")
    def test_excel_creates_file(self, tmp_path):
        rows = self._make_rows()
        output = str(tmp_path / "test.xlsx")
        generate_excel(rows, rows, output, "develop", "feature")
        assert os.path.exists(output)

    @pytest.mark.skipif(not pytest.importorskip("openpyxl"), reason="openpyxl required")
    def test_excel_has_three_sheets(self, tmp_path):
        from openpyxl import load_workbook
        rows = self._make_rows()
        output = str(tmp_path / "test.xlsx")
        generate_excel(rows, rows, output, "develop", "feature")
        wb = load_workbook(output)
        assert 'Summary' in wb.sheetnames
        assert 'GEMM' in wb.sheetnames
        assert 'Conv' in wb.sheetnames
        wb.close()

    @pytest.mark.skipif(not pytest.importorskip("openpyxl"), reason="openpyxl required")
    def test_excel_correct_row_count(self, tmp_path):
        from openpyxl import load_workbook
        gemm_rows = self._make_rows()
        conv_rows = self._make_rows()[:2]
        output = str(tmp_path / "test.xlsx")
        generate_excel(gemm_rows, conv_rows, output, "develop", "feature")
        wb = load_workbook(output)
        assert wb['GEMM'].max_row == 4  # 1 header + 3 data
        assert wb['Conv'].max_row == 3  # 1 header + 2 data
        wb.close()

    @pytest.mark.skipif(not pytest.importorskip("openpyxl"), reason="openpyxl required")
    def test_excel_header_content(self, tmp_path):
        from openpyxl import load_workbook
        rows = self._make_rows()
        output = str(tmp_path / "test.xlsx")
        generate_excel(rows, rows, output, "develop", "feature")
        wb = load_workbook(output)
        ws = wb['GEMM']
        assert ws.cell(row=1, column=1).value == 'Problem Config'
        assert 'develop' in ws.cell(row=1, column=2).value
        assert 'feature' in ws.cell(row=1, column=3).value
        assert 'Diff' in ws.cell(row=1, column=8).value
        wb.close()

    @pytest.mark.skipif(not pytest.importorskip("openpyxl"), reason="openpyxl required")
    def test_excel_data_content(self, tmp_path):
        from openpyxl import load_workbook
        rows = self._make_rows()
        output = str(tmp_path / "test.xlsx")
        generate_excel(rows, rows, output, "develop", "feature")
        wb = load_workbook(output)
        ws = wb['GEMM']
        assert ws.cell(row=2, column=1).value == 'cfg_a'
        assert ws.cell(row=2, column=4).value == 100.0
        assert ws.cell(row=2, column=5).value == 110.0
        assert ws.cell(row=2, column=6).value == 4
        assert ws.cell(row=2, column=7).value == 16
        assert abs(ws.cell(row=2, column=8).value - 10.0) < 0.01
        wb.close()

    @pytest.mark.skipif(not pytest.importorskip("openpyxl"), reason="openpyxl required")
    def test_excel_header_bold(self, tmp_path):
        from openpyxl import load_workbook
        rows = self._make_rows()
        output = str(tmp_path / "test.xlsx")
        generate_excel(rows, rows, output, "develop", "feature")
        wb = load_workbook(output)
        ws = wb['GEMM']
        for col in range(1, 9):
            assert ws.cell(row=1, column=col).font.bold
        wb.close()

    @pytest.mark.skipif(not pytest.importorskip("openpyxl"), reason="openpyxl required")
    def test_validation_passes(self, tmp_path):
        rows = self._make_rows()
        output = str(tmp_path / "test.xlsx")
        generate_excel(rows, rows, output, "develop", "feature")
        assert validate_excel(output, rows, rows, num_checks=3)

    @pytest.mark.skipif(not pytest.importorskip("openpyxl"), reason="openpyxl required")
    def test_summary_sheet_metadata(self, tmp_path):
        from openpyxl import load_workbook
        rows = self._make_rows()
        output = str(tmp_path / "test.xlsx")
        metadata = ExcelMetadata(
            base_branch="develop",
            feature_branch="swapOperands2",
            base_commit="abc1234",
            feature_commit="def5678",
            hostname="test-server-01",
            gpu_arch="gfx950",
        )
        generate_excel(rows, rows, output, "develop", "swapOperands2", metadata=metadata)
        wb = load_workbook(output)
        ws = wb['Summary']

        # Read all property-value pairs into a dict
        props = {}
        for row_idx in range(2, ws.max_row + 1):
            prop = ws.cell(row=row_idx, column=1).value
            val = ws.cell(row=row_idx, column=2).value
            if prop:
                props[prop] = val

        assert props['Hostname'] == 'test-server-01'
        assert props['GPU Arch'] == 'gfx950'
        assert props['Base Branch'] == 'develop'
        assert props['Base Commit'] == 'abc1234'
        assert props['Feature Branch'] == 'swapOperands2'
        assert props['Feature Commit'] == 'def5678'
        assert props['GEMM Configs'] == 3
        assert props['Conv Configs'] == 3
        wb.close()


class TestSummarySheetMetadataDetails:
    """Detailed tests for each metadata field in the Summary sheet."""

    def _make_rows(self):
        return [
            ComparisonRow("cfg_a", "pc_b", "pc_f", 100.0, 110.0, 4, 16, 10.0),
            ComparisonRow("cfg_b", "pc_b2", "pc_f2", 50.0, 45.0, 8, 8, -10.0),
        ]

    def _read_summary_props(self, output_path):
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb['Summary']
        props = {}
        for row_idx in range(2, ws.max_row + 1):
            prop = ws.cell(row=row_idx, column=1).value
            val = ws.cell(row=row_idx, column=2).value
            if prop:
                props[prop] = val
        wb.close()
        return props

    @pytest.mark.skipif(not pytest.importorskip("openpyxl"), reason="openpyxl required")
    def test_hostname_present(self, tmp_path):
        rows = self._make_rows()
        output = str(tmp_path / "test.xlsx")
        metadata = ExcelMetadata(hostname="my-perf-server-42")
        generate_excel(rows, rows, output, "base", "feat", metadata=metadata)
        props = self._read_summary_props(output)
        assert props['Hostname'] == 'my-perf-server-42'

    @pytest.mark.skipif(not pytest.importorskip("openpyxl"), reason="openpyxl required")
    def test_hostname_empty_shows_placeholder(self, tmp_path):
        rows = self._make_rows()
        output = str(tmp_path / "test.xlsx")
        metadata = ExcelMetadata(hostname="")
        generate_excel(rows, rows, output, "base", "feat", metadata=metadata)
        props = self._read_summary_props(output)
        assert props['Hostname'] == '(not provided)'

    @pytest.mark.skipif(not pytest.importorskip("openpyxl"), reason="openpyxl required")
    def test_gpu_arch_present(self, tmp_path):
        rows = self._make_rows()
        output = str(tmp_path / "test.xlsx")
        metadata = ExcelMetadata(gpu_arch="gfx942")
        generate_excel(rows, rows, output, "base", "feat", metadata=metadata)
        props = self._read_summary_props(output)
        assert props['GPU Arch'] == 'gfx942'

    @pytest.mark.skipif(not pytest.importorskip("openpyxl"), reason="openpyxl required")
    def test_gpu_arch_empty_shows_placeholder(self, tmp_path):
        rows = self._make_rows()
        output = str(tmp_path / "test.xlsx")
        metadata = ExcelMetadata(gpu_arch="")
        generate_excel(rows, rows, output, "base", "feat", metadata=metadata)
        props = self._read_summary_props(output)
        assert props['GPU Arch'] == '(not provided)'

    @pytest.mark.skipif(not pytest.importorskip("openpyxl"), reason="openpyxl required")
    def test_base_commit_present(self, tmp_path):
        rows = self._make_rows()
        output = str(tmp_path / "test.xlsx")
        metadata = ExcelMetadata(base_commit="031a50a")
        generate_excel(rows, rows, output, "base", "feat", metadata=metadata)
        props = self._read_summary_props(output)
        assert props['Base Commit'] == '031a50a'

    @pytest.mark.skipif(not pytest.importorskip("openpyxl"), reason="openpyxl required")
    def test_feature_commit_present(self, tmp_path):
        rows = self._make_rows()
        output = str(tmp_path / "test.xlsx")
        metadata = ExcelMetadata(feature_commit="f7e3b21")
        generate_excel(rows, rows, output, "base", "feat", metadata=metadata)
        props = self._read_summary_props(output)
        assert props['Feature Commit'] == 'f7e3b21'

    @pytest.mark.skipif(not pytest.importorskip("openpyxl"), reason="openpyxl required")
    def test_commits_empty_show_placeholder(self, tmp_path):
        rows = self._make_rows()
        output = str(tmp_path / "test.xlsx")
        metadata = ExcelMetadata(base_commit="", feature_commit="")
        generate_excel(rows, rows, output, "base", "feat", metadata=metadata)
        props = self._read_summary_props(output)
        assert props['Base Commit'] == '(not provided)'
        assert props['Feature Commit'] == '(not provided)'

    @pytest.mark.skipif(not pytest.importorskip("openpyxl"), reason="openpyxl required")
    def test_date_is_utc_formatted(self, tmp_path):
        import re
        rows = self._make_rows()
        output = str(tmp_path / "test.xlsx")
        metadata = ExcelMetadata()
        generate_excel(rows, rows, output, "base", "feat", metadata=metadata)
        props = self._read_summary_props(output)
        date_val = props['Date']
        assert re.match(r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2} UTC', date_val), \
            f"Date '{date_val}' doesn't match expected format"

    @pytest.mark.skipif(not pytest.importorskip("openpyxl"), reason="openpyxl required")
    def test_branch_names_present(self, tmp_path):
        rows = self._make_rows()
        output = str(tmp_path / "test.xlsx")
        metadata = ExcelMetadata(base_branch="main", feature_branch="my-feature")
        generate_excel(rows, rows, output, "main", "my-feature", metadata=metadata)
        props = self._read_summary_props(output)
        assert props['Base Branch'] == 'main'
        assert props['Feature Branch'] == 'my-feature'

    @pytest.mark.skipif(not pytest.importorskip("openpyxl"), reason="openpyxl required")
    def test_all_metadata_fields_together(self, tmp_path):
        rows = self._make_rows()
        output = str(tmp_path / "test.xlsx")
        metadata = ExcelMetadata(
            base_branch="develop",
            feature_branch="swapOperands2",
            base_commit="aaa1111",
            feature_commit="bbb2222",
            hostname="cv-MI350",
            gpu_arch="gfx950",
        )
        generate_excel(rows, rows, output, "develop", "swapOperands2", metadata=metadata)
        props = self._read_summary_props(output)
        assert props['Hostname'] == 'cv-MI350'
        assert props['GPU Arch'] == 'gfx950'
        assert props['Base Branch'] == 'develop'
        assert props['Base Commit'] == 'aaa1111'
        assert props['Feature Branch'] == 'swapOperands2'
        assert props['Feature Commit'] == 'bbb2222'
        assert 'Date' in props
        assert 'UTC' in props['Date']


# ---------------------------------------------------------------------------
# Hardware Detection Tests
# ---------------------------------------------------------------------------


class TestHardwareDetection:

    def test_parse_cpu_list_simple(self):
        assert _parse_cpu_list("0-3") == [0, 1, 2, 3]

    def test_parse_cpu_list_multi_range(self):
        result = _parse_cpu_list("0-3,8-11")
        assert result == [0, 1, 2, 3, 8, 9, 10, 11]

    def test_parse_cpu_list_single(self):
        assert _parse_cpu_list("5") == [5]

    def test_parse_cpu_list_mixed(self):
        result = _parse_cpu_list("0-2,5,8-9")
        assert result == [0, 1, 2, 5, 8, 9]

    def test_parse_cpu_list_empty(self):
        assert _parse_cpu_list("") == []

    @mock.patch('subprocess.check_output')
    def test_detect_gpus_from_rocm_smi(self, mock_check):
        mock_json = {
            "card0": {
                "Card SKU": "MI300X",
                "(Topology) Numa Node": "0"
            },
            "card1": {
                "Card SKU": "MI300X",
                "(Topology) Numa Node": "1"
            },
            "card2": {
                "Card SKU": "MI300X",
                "(Topology) Numa Node": "2"
            },
            "card3": {
                "Card SKU": "MI300X",
                "(Topology) Numa Node": "3"
            },
        }
        import json
        mock_check.return_value = json.dumps(mock_json)
        gpus = detect_gpus_rocm_smi()
        assert gpus == [0, 1, 2, 3]

    @mock.patch('subprocess.check_output', side_effect=Exception("not found"))
    def test_detect_gpus_fallback(self, mock_check):
        gpus = detect_gpus_rocm_smi()
        assert gpus == [0]


# ---------------------------------------------------------------------------
# Conv Flag Parsing Tests
# ---------------------------------------------------------------------------


class TestParseConvFlags:

    def test_basic(self):
        argv = ["-F", "1", "-f", "NCHW", "-n", "1"]
        result = _parse_conv_flags(argv)
        assert result['F'] == '1'
        assert result['f'] == 'NCHW'
        assert result['n'] == '1'

    def test_full_conv_flags(self):
        argv = [
            "-F", "1", "-f", "GNC01", "-I", "NGC01", "-O", "NGC01", "-n", "1", "-c", "256", "-H",
            "14", "-W", "14", "-k", "1024", "-y", "1", "-x", "1", "-p", "0", "-q", "0", "-u", "1",
            "-v", "1", "-l", "1", "-j", "1", "-g", "1"
        ]
        result = _parse_conv_flags(argv)
        assert result['F'] == '1'
        assert result['c'] == '256'
        assert result['k'] == '1024'
        assert result['g'] == '1'


# ---------------------------------------------------------------------------
# End-to-end Integration Test (with synthetic data)
# ---------------------------------------------------------------------------


class TestEndToEnd:
    """Test the full pipeline with synthetic data (no GPU needed)."""

    @pytest.mark.skipif(not pytest.importorskip("openpyxl"), reason="openpyxl required")
    def test_full_pipeline(self, tmp_path):
        from openpyxl import load_workbook

        gemm_tsv_content = (
            "# commit: abc\n"
            "# arch\tnumCUs\tnumChiplets\ttestVector\tperfConfig\n"
            "gfx950\t256\t8\t-t f32 -out_datatype f32 -transA false -transB false -g 1 -m 100 -n 200 -k 300\tpc_1\n"
            "gfx950\t256\t8\t-t f16 -out_datatype f16 -transA true -transB false -g 1 -m 50 -n 60 -k 70\tpc_2\n"
        )

        gemm_csv_headers = [
            'DataType', 'OutDataType', 'Chip', 'numCU', 'numChiplets', 'TransA', 'TransB', 'G', 'M',
            'K', 'N', 'ScaledGemm', 'ScaleADtype', 'ScaleBDtype', 'TransScaleA', 'TransScaleB',
            'PerfConfig', 'LDSBankConflict', 'TFlops'
        ]
        base_gemm_csv_rows = [
            [
                'f32', 'f32', 'gfx950', 256, 8, 'false', 'false', 1, 100, 300, 200, 'False', '', '',
                'False', 'False', 'pc_1', 0, 50.0
            ],
            [
                'f16', 'f16', 'gfx950', 256, 8, 'true', 'false', 1, 50, 70, 60, 'False', '', '',
                'False', 'False', 'pc_2', 0, 30.0
            ],
        ]
        feat_gemm_csv_rows = [
            [
                'f32', 'f32', 'gfx950', 256, 8, 'false', 'false', 1, 100, 300, 200, 'False', '', '',
                'False', 'False', 'pc_1f', 0, 55.0
            ],
            [
                'f16', 'f16', 'gfx950', 256, 8, 'true', 'false', 1, 50, 70, 60, 'False', '', '',
                'False', 'False', 'pc_2f', 0, 28.0
            ],
        ]

        conv_tsv_content = (
            "# commit: abc\n"
            "# arch\tnumCUs\tnumChiplets\ttestVector\tperfConfig\n"
            "gfx950\t256\t8\tconv -F 1 -f GNC01 -I NGC01 -O NGC01 -n 1 -c 64 -H 56 -W 56 -k 64 -y 3 -x 3 -p 1 -q 1 -u 1 -v 1 -l 1 -j 1 -g 1\tpc_conv1\n"
        )

        conv_csv_headers = [
            'Direction', 'DataType', 'Chip', 'numCU', 'numChiplets', 'FilterLayout', 'InputLayout',
            'OutputLayout', 'N', 'C', 'H', 'W', 'K', 'Y', 'X', 'DilationH', 'DilationW', 'StrideH',
            'StrideW', 'PaddingH', 'PaddingW', 'PerfConfig', 'LDSBankConflict', 'TFlops'
        ]
        base_conv_csv_rows = [
            [
                'fwd', 'f32', 'gfx950', 256, 8, 'gkcyx', 'ngchw', 'ngkhw', 1, 64, 56, 56, 64, 3, 3,
                1, 1, 1, 1, 1, 1, 'pc_conv1', 0, 20.0
            ],
        ]
        feat_conv_csv_rows = [
            [
                'fwd', 'f32', 'gfx950', 256, 8, 'gkcyx', 'ngchw', 'ngkhw', 1, 64, 56, 56, 64, 3, 3,
                1, 1, 1, 1, 1, 1, 'pc_conv1f', 0, 22.0
            ],
        ]

        # Write files
        base_gemm_tsv = tmp_path / "base_gemm.tsv"
        base_gemm_tsv.write_text(gemm_tsv_content)
        feat_gemm_tsv = tmp_path / "feat_gemm.tsv"
        feat_gemm_tsv.write_text(gemm_tsv_content.replace("pc_1", "pc_1f").replace("pc_2", "pc_2f"))

        base_gemm_csv = tmp_path / "base_gemm.csv"
        pd.DataFrame(base_gemm_csv_rows, columns=gemm_csv_headers).to_csv(str(base_gemm_csv),
                                                                          index=False)
        feat_gemm_csv = tmp_path / "feat_gemm.csv"
        pd.DataFrame(feat_gemm_csv_rows, columns=gemm_csv_headers).to_csv(str(feat_gemm_csv),
                                                                          index=False)

        base_conv_tsv = tmp_path / "base_conv.tsv"
        base_conv_tsv.write_text(conv_tsv_content)
        feat_conv_tsv = tmp_path / "feat_conv.tsv"
        feat_conv_tsv.write_text(conv_tsv_content.replace("pc_conv1", "pc_conv1f"))

        base_conv_csv = tmp_path / "base_conv.csv"
        pd.DataFrame(base_conv_csv_rows, columns=conv_csv_headers).to_csv(str(base_conv_csv),
                                                                          index=False)
        feat_conv_csv = tmp_path / "feat_conv.csv"
        pd.DataFrame(feat_conv_csv_rows, columns=conv_csv_headers).to_csv(str(feat_conv_csv),
                                                                          index=False)

        output_xlsx = tmp_path / "comparison.xlsx"

        # Parse
        base_gemm_t = parse_tuning_tsv(str(base_gemm_tsv))
        feat_gemm_t = parse_tuning_tsv(str(feat_gemm_tsv))
        base_gemm_p = parse_perf_csv(str(base_gemm_csv), 'gemm')
        feat_gemm_p = parse_perf_csv(str(feat_gemm_csv), 'gemm')

        base_conv_t = parse_tuning_tsv(str(base_conv_tsv))
        feat_conv_t = parse_tuning_tsv(str(feat_conv_tsv))
        base_conv_p = parse_perf_csv(str(base_conv_csv), 'conv')
        feat_conv_p = parse_perf_csv(str(feat_conv_csv), 'conv')

        # Merge (skip ISA)
        gemm_rows = sort_comparison_rows(
            merge_results(base_gemm_t, feat_gemm_t, base_gemm_p, feat_gemm_p, {}, {}))
        conv_rows = sort_comparison_rows(
            merge_results(base_conv_t, feat_conv_t, base_conv_p, feat_conv_p, {}, {}))

        # Excel
        generate_excel(gemm_rows, conv_rows, str(output_xlsx), "develop", "feature")
        assert output_xlsx.exists()

        # Validate
        wb = load_workbook(str(output_xlsx))
        gemm_ws = wb['GEMM']
        conv_ws = wb['Conv']

        assert gemm_ws.max_row >= 3  # header + 2 data rows
        assert conv_ws.max_row >= 2  # header + 1 data row

        # Check GEMM data: cfg with +10% improvement should come first
        gemm_diffs = []
        for row_idx in range(2, gemm_ws.max_row + 1):
            diff = gemm_ws.cell(row=row_idx, column=8).value
            if diff is not None:
                gemm_diffs.append(float(diff))

        assert len(gemm_diffs) == 2
        assert gemm_diffs[0] >= gemm_diffs[1], "Diffs should be sorted descending"

        wb.close()

    @pytest.mark.skipif(not pytest.importorskip("openpyxl"), reason="openpyxl required")
    def test_independent_validation_against_sources(self, tmp_path):
        """Test that validate_excel_against_sources correctly verifies
        Excel content against the original TSV/CSV source files."""
        gemm_tsv_content = (
            "# commit: abc\n"
            "# arch\tnumCUs\tnumChiplets\ttestVector\tperfConfig\n"
            "gfx950\t256\t8\t-t f32 -out_datatype f32 -transA false -transB false -g 1 -m 100 -n 200 -k 300\tpc_base\n"
        )
        feat_gemm_tsv_content = (
            "# commit: def\n"
            "# arch\tnumCUs\tnumChiplets\ttestVector\tperfConfig\n"
            "gfx950\t256\t8\t-t f32 -out_datatype f32 -transA false -transB false -g 1 -m 100 -n 200 -k 300\tpc_feat\n"
        )

        gemm_csv_headers = [
            'DataType', 'OutDataType', 'Chip', 'numCU', 'numChiplets', 'TransA', 'TransB', 'G', 'M',
            'K', 'N', 'ScaledGemm', 'ScaleADtype', 'ScaleBDtype', 'TransScaleA', 'TransScaleB',
            'PerfConfig', 'LDSBankConflict', 'TFlops'
        ]
        base_csv_rows = [[
            'f32', 'f32', 'gfx950', 256, 8, 'false', 'false', 1, 100, 300, 200, 'False', '', '',
            'False', 'False', 'pc_base', 0, 50.0
        ]]
        feat_csv_rows = [[
            'f32', 'f32', 'gfx950', 256, 8, 'false', 'false', 1, 100, 300, 200, 'False', '', '',
            'False', 'False', 'pc_feat', 0, 55.0
        ]]

        conv_tsv_content = (
            "# commit: abc\n"
            "# arch\tnumCUs\tnumChiplets\ttestVector\tperfConfig\n"
            "gfx950\t256\t8\tconv -F 1 -f GNC01 -I NGC01 -O NGC01 -n 1 -c 64 -H 56 -W 56 -k 64 -y 3 -x 3 -p 1 -q 1 -u 1 -v 1 -l 1 -j 1 -g 1\tpc_cv\n"
        )
        conv_csv_headers = [
            'Direction', 'DataType', 'Chip', 'numCU', 'numChiplets', 'FilterLayout', 'InputLayout',
            'OutputLayout', 'N', 'C', 'H', 'W', 'K', 'Y', 'X', 'DilationH', 'DilationW', 'StrideH',
            'StrideW', 'PaddingH', 'PaddingW', 'PerfConfig', 'LDSBankConflict', 'TFlops'
        ]
        conv_csv_rows = [[
            'fwd', 'f32', 'gfx950', 256, 8, 'gkcyx', 'ngchw', 'ngkhw', 1, 64, 56, 56, 64, 3, 3, 1,
            1, 1, 1, 1, 1, 'pc_cv', 0, 20.0
        ]]

        # Write all source files
        b_g_tsv = tmp_path / "b_gemm.tsv"
        b_g_tsv.write_text(gemm_tsv_content)
        f_g_tsv = tmp_path / "f_gemm.tsv"
        f_g_tsv.write_text(feat_gemm_tsv_content)

        b_g_csv = tmp_path / "b_gemm.csv"
        pd.DataFrame(base_csv_rows, columns=gemm_csv_headers).to_csv(str(b_g_csv), index=False)
        f_g_csv = tmp_path / "f_gemm.csv"
        pd.DataFrame(feat_csv_rows, columns=gemm_csv_headers).to_csv(str(f_g_csv), index=False)

        b_c_tsv = tmp_path / "b_conv.tsv"
        b_c_tsv.write_text(conv_tsv_content)
        f_c_tsv = tmp_path / "f_conv.tsv"
        f_c_tsv.write_text(conv_tsv_content.replace("pc_cv", "pc_cvf"))

        b_c_csv = tmp_path / "b_conv.csv"
        pd.DataFrame(conv_csv_rows, columns=conv_csv_headers).to_csv(str(b_c_csv), index=False)
        f_c_csv = tmp_path / "f_conv.csv"
        feat_conv_rows = [list(conv_csv_rows[0])]
        feat_conv_rows[0][-1] = 22.0  # different TFlops
        feat_conv_rows[0][-3] = 'pc_cvf'
        pd.DataFrame(feat_conv_rows, columns=conv_csv_headers).to_csv(str(f_c_csv), index=False)

        # Build Excel through the normal pipeline
        base_gemm_t = parse_tuning_tsv(str(b_g_tsv))
        feat_gemm_t = parse_tuning_tsv(str(f_g_tsv))
        base_gemm_p = parse_perf_csv(str(b_g_csv), 'gemm')
        feat_gemm_p = parse_perf_csv(str(f_g_csv), 'gemm')
        base_conv_t = parse_tuning_tsv(str(b_c_tsv))
        feat_conv_t = parse_tuning_tsv(str(f_c_tsv))
        base_conv_p = parse_perf_csv(str(b_c_csv), 'conv')
        feat_conv_p = parse_perf_csv(str(f_c_csv), 'conv')

        gemm_rows = sort_comparison_rows(
            merge_results(base_gemm_t, feat_gemm_t, base_gemm_p, feat_gemm_p, {}, {}))
        conv_rows = sort_comparison_rows(
            merge_results(base_conv_t, feat_conv_t, base_conv_p, feat_conv_p, {}, {}))

        output = str(tmp_path / "out.xlsx")
        generate_excel(gemm_rows, conv_rows, output, "develop", "feature")

        # Now run the independent validation
        ok = validate_excel_against_sources(
            excel_path=output,
            base_gemm_tsv=str(b_g_tsv),
            base_gemm_csv=str(b_g_csv),
            feat_gemm_tsv=str(f_g_tsv),
            feat_gemm_csv=str(f_g_csv),
            base_conv_tsv=str(b_c_tsv),
            base_conv_csv=str(b_c_csv),
            feat_conv_tsv=str(f_c_tsv),
            feat_conv_csv=str(f_c_csv),
            base_branch="develop",
            feature_branch="feature",
            num_checks=5,
        )
        assert ok, "Independent validation should pass for correctly generated Excel"
