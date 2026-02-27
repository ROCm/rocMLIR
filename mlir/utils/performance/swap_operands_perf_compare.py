#!/usr/bin/env python3
"""A/B performance comparison between two rocMLIR branches.

Parses tuning DB (TSV) and perf results (CSV) from two branches,
extracts ISA global_store vector widths, and generates a comparison
Excel report with diff analysis.

Usage (standalone, after tuning/perf runs are complete):
    python3 swap_operands_perf_compare.py \
        --feature-branch swapOperands2 --base-branch develop \
        --feature-build-dir ./build --base-build-dir ./build_develop \
        --feature-gemm-tsv results/swapOperands2_gemm.tsv \
        --base-gemm-tsv results/develop_gemm.tsv \
        --feature-gemm-csv results/swapOperands2_gemm_perf.csv \
        --base-gemm-csv results/develop_gemm_perf.csv \
        --feature-conv-tsv results/swapOperands2_conv.tsv \
        --base-conv-tsv results/develop_conv.tsv \
        --feature-conv-csv results/swapOperands2_conv_perf.csv \
        --base-conv-csv results/develop_conv_perf.csv \
        --gemm-configs mlir/utils/performance/configs/tier1-gemm-configs \
        --conv-configs mlir/utils/performance/configs/tier1-conv-configs \
        --output comparison.xlsx
"""

import argparse
import csv
import json
import os
import random
import re
import subprocess
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill, numbers
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ---------------------------------------------------------------------------
# Hardware detection
# ---------------------------------------------------------------------------


def detect_gpus_rocm_smi() -> List[int]:
    """Detect available GPU IDs via rocm-smi."""
    try:
        output = subprocess.check_output(
            ["rocm-smi", "--showproductname", "--showtoponuma", "--json"],
            text=True,
            timeout=10,
            stderr=subprocess.DEVNULL)
        data = json.loads(output)
        gpu_ids = sorted(int(k.replace("card", "")) for k in data if k.startswith("card"))
        return gpu_ids if gpu_ids else list(range(1))
    except Exception:
        return list(range(1))


def detect_num_cpus() -> int:
    """Detect available CPU count via NUMA topology or os.cpu_count()."""
    numa_base = "/sys/devices/system/node"
    total = 0
    if os.path.exists(numa_base):
        for entry in os.listdir(numa_base):
            if entry.startswith("node") and entry[4:].isdigit():
                cpulist_path = os.path.join(numa_base, entry, "cpulist")
                try:
                    with open(cpulist_path, 'r') as f:
                        total += len(_parse_cpu_list(f.read()))
                except OSError:
                    pass
    return total if total > 0 else (os.cpu_count() or 1)


def _parse_cpu_list(cpu_list_str: str) -> List[int]:
    """Parse CPU list string like '0-55,112-167' into list of CPU IDs."""
    cpus: List[int] = []
    for part in cpu_list_str.strip().split(','):
        if not part:
            continue
        if '-' in part:
            lo, hi = part.split('-', 1)
            cpus.extend(range(int(lo), int(hi) + 1))
        else:
            cpus.append(int(part))
    return cpus


def detect_gpu_arch() -> str:
    """Return a human-readable GPU arch string."""
    try:
        output = subprocess.check_output(["rocm-smi", "--showproductname", "--json"],
                                         text=True,
                                         timeout=10,
                                         stderr=subprocess.DEVNULL)
        data = json.loads(output)
        for k, v in data.items():
            if k.startswith("card"):
                return v.get("Card SKU", "unknown")
        return "unknown"
    except Exception:
        return "unknown"


def print_hardware_info(gpu_ids: List[int], num_cpus: int) -> None:
    arch = detect_gpu_arch()
    print(f"Detected hardware:")
    print(f"  GPUs: {gpu_ids} ({len(gpu_ids)} GPUs, {arch})")
    print(f"  CPUs: {num_cpus}")


# ---------------------------------------------------------------------------
# TSV (tuning DB) parsing
# ---------------------------------------------------------------------------


@dataclass
class TuningEntry:
    arch: str
    num_cu: int
    num_chiplets: int
    test_vector: str
    perf_config: str


def parse_tuning_tsv(path: str) -> Dict[str, TuningEntry]:
    """Parse a tuningRunner TSV and return a dict keyed by test_vector."""
    entries: Dict[str, TuningEntry] = {}
    with open(path, 'r') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            parts = line.split('\t')
            if len(parts) < 5:
                # v2 format: arch, numCU, testVector, perfConfig
                if len(parts) == 4:
                    arch, num_cu, tv, pc = parts
                    entry = TuningEntry(arch, int(num_cu), 0, tv.strip(), pc.strip())
                elif len(parts) == 3:
                    arch, tv, pc = parts
                    entry = TuningEntry(arch, 0, 0, tv.strip(), pc.strip())
                else:
                    continue
            else:
                arch, num_cu, num_chiplets, tv, pc = parts[:5]
                entry = TuningEntry(arch, int(num_cu), int(num_chiplets), tv.strip(), pc.strip())
            entries[entry.test_vector] = entry
    return entries


# ---------------------------------------------------------------------------
# CSV (perf results) parsing
# ---------------------------------------------------------------------------

GEMM_CSV_COLUMNS = [
    'DataType', 'OutDataType', 'Chip', 'numCU', 'numChiplets', 'TransA', 'TransB', 'G', 'M', 'K',
    'N', 'ScaledGemm', 'ScaleADtype', 'ScaleBDtype', 'TransScaleA', 'TransScaleB', 'PerfConfig',
    'LDSBankConflict', 'TFlops'
]

CONV_CSV_COLUMNS = [
    'Direction', 'DataType', 'Chip', 'numCU', 'numChiplets', 'FilterLayout', 'InputLayout',
    'OutputLayout', 'N', 'C', 'H', 'W', 'K', 'Y', 'X', 'DilationH', 'DilationW', 'StrideH',
    'StrideW', 'PaddingH', 'PaddingW', 'PerfConfig', 'LDSBankConflict', 'TFlops'
]


def gemm_row_to_test_vector(row: dict) -> str:
    """Reconstruct the tuning test_vector key from a GEMM CSV row."""
    trans_a = str(row['TransA']).lower()
    trans_b = str(row['TransB']).lower()
    result = (f"-t {row['DataType']} -out_datatype {row['OutDataType']} "
              f"-transA {trans_a} -transB {trans_b} "
              f"-g {row['G']} -m {row['M']} -n {row['N']} -k {row['K']}")
    if str(row.get('ScaledGemm', 'False')).lower() == 'true':
        result += " -scaledGemm"
    if row.get('ScaleADtype') and str(row['ScaleADtype']) not in ('', 'None', 'nan'):
        result += f" -scale_a_dtype {row['ScaleADtype']}"
    if row.get('ScaleBDtype') and str(row['ScaleBDtype']) not in ('', 'None', 'nan'):
        result += f" -scale_b_dtype {row['ScaleBDtype']}"
    if str(row.get('TransScaleA', 'False')).lower() == 'true':
        result += f" -transScaleA {str(row['TransScaleA']).lower()}"
    if str(row.get('TransScaleB', 'False')).lower() == 'true':
        result += f" -transScaleB {str(row['TransScaleB']).lower()}"
    return result


CONV_DTYPE_PREFIX = {
    'f32': 'conv',
    'f16': 'convfp16',
    'bf16': 'convbfp16',
    'i8': 'convint8',
    'fp8': 'convfp8',
    'fp8_fp8': 'convfp8_fp8',
    'fp8_bf8': 'convfp8_bf8',
    'bf8_fp8': 'convbf8_fp8',
    'bf8_bf8': 'convbf8_bf8',
}

CONV_DIR_MAP = {'fwd': '1', 'bwd': '2', 'wrw': '4'}

FILTER_LAYOUT_INV = {'k': 'N', 'c': 'C', 'y': 'H', 'x': 'W', 'g': 'G', '0': '0', '1': '1'}
OUTPUT_LAYOUT_INV = {'n': 'N', 'k': 'C', 'h': 'H', 'w': 'W', 'g': 'G', '0': '0', '1': '1'}


def _inverse_filter_layout(layout: str) -> str:
    return "".join(FILTER_LAYOUT_INV.get(c, c) for c in layout)


def _inverse_output_layout(layout: str) -> str:
    return "".join(OUTPUT_LAYOUT_INV.get(c, c) for c in layout)


def conv_row_to_test_vector(row: dict) -> str:
    """Reconstruct the tuning test_vector key from a Conv CSV row."""
    dtype = str(row['DataType'])
    prefix = CONV_DTYPE_PREFIX.get(dtype, 'conv')
    direction = str(row['Direction'])
    dir_num = CONV_DIR_MAP.get(direction, '1')
    fil = _inverse_filter_layout(str(row['FilterLayout']))
    inp = str(row['InputLayout']).upper()
    out = _inverse_output_layout(str(row['OutputLayout']))
    return (f"{prefix} -F {dir_num} -f {fil} -I {inp} -O {out} "
            f"-n {row['N']} -c {row['C']} -H {row['H']} -W {row['W']} "
            f"-k {row['K']} -y {row['Y']} -x {row['X']} "
            f"-p {row['PaddingH']} -q {row['PaddingW']} "
            f"-u {row['StrideH']} -v {row['StrideW']} "
            f"-l {row['DilationH']} -j {row['DilationW']} -m conv -g 1 -t 1")


@dataclass
class PerfEntry:
    test_vector: str
    perf_config: str
    tflops: float


def parse_perf_csv(path: str, operation: str) -> Dict[str, PerfEntry]:
    """Parse a perfRunner CSV and return dict keyed by test_vector."""
    entries: Dict[str, PerfEntry] = {}
    df = pd.read_csv(path)
    for _, row in df.iterrows():
        if operation == 'gemm':
            tv = gemm_row_to_test_vector(row)
        else:
            tv = conv_row_to_test_vector(row)
        pc = str(row.get('PerfConfig', ''))
        tflops = float(row.get('TFlops', float('nan')))
        entries[tv] = PerfEntry(tv, pc, tflops)
    return entries


# ---------------------------------------------------------------------------
# ISA global_store width extraction
# ---------------------------------------------------------------------------

GLOBAL_STORE_RE = re.compile(r'global_store_(\w+)')

STORE_WIDTH_MAP = {
    'byte': 1,
    'byte_d16_hi': 1,
    'short': 2,
    'short_d16_hi': 2,
    'dword': 4,
    'dwordx2': 8,
    'dwordx3': 12,
    'dwordx4': 16,
    'b8': 1,
    'b16': 2,
    'b32': 4,
    'b64': 8,
    'b96': 12,
    'b128': 16,
}


def extract_global_store_width_from_isa(isa_text: str) -> int:
    """Extract the maximum global_store vector width (in bytes) from ISA text."""
    max_width = 0
    for match in GLOBAL_STORE_RE.finditer(isa_text):
        suffix = match.group(1)
        width = STORE_WIDTH_MAP.get(suffix, 0)
        max_width = max(max_width, width)
    return max_width


def get_isa_for_config(build_dir: str, operation: str, test_vector: str, perf_config: str,
                       arch: str, num_cu: int, num_chiplets: int) -> str:
    """Run rocmlir-gen | rocmlir-driver -c --debug-only=serialize-to-isa
    and return the ISA text (from stderr)."""
    rocmlir_gen = os.path.join(build_dir, 'bin', 'rocmlir-gen')
    rocmlir_driver = os.path.join(build_dir, 'bin', 'rocmlir-driver')

    if not os.path.exists(rocmlir_gen) or not os.path.exists(rocmlir_driver):
        return ""

    if operation == 'gemm':
        gen_args = (f"-operation gemm {test_vector} --arch {arch} "
                    f"--num_cu {num_cu} --num_chiplets {num_chiplets} "
                    f"--perf_config={perf_config}")
    else:
        parts = test_vector.split()
        dtype_prefix = parts[0] if parts else 'conv'
        dtype_map = {
            'conv': 'f32',
            'convfp16': 'f16',
            'convbfp16': 'bf16',
            'convint8': 'i8',
            'convfp8': 'fp8',
            'convfp8_fp8': 'fp8_fp8',
            'convfp8_bf8': 'fp8_bf8',
            'convbf8_fp8': 'bf8_fp8',
            'convbf8_bf8': 'bf8_bf8',
        }
        dtype = dtype_map.get(dtype_prefix, 'f32')
        flags = _parse_conv_flags(parts[1:])
        direction_map = {'1': 'conv', '2': 'conv_bwd_data', '4': 'conv_bwd_weight'}
        op_name = direction_map.get(flags.get('F', '1'), 'conv')
        gen_args = (f"--operation {op_name} -t {dtype} --arch {arch} "
                    f"--num_cu {num_cu} --num_chiplets {num_chiplets} "
                    f"--fil_layout {flags.get('f', 'NCHW').lower()} "
                    f"--in_layout {flags.get('I', 'NCHW').lower()} "
                    f"--out_layout {flags.get('O', 'NCHW').lower()} "
                    f"--batchsize {flags.get('n', '1')} "
                    f"--in_channels {flags.get('c', '1')} "
                    f"--in_h {flags.get('H', '1')} --in_w {flags.get('W', '1')} "
                    f"--out_channels {flags.get('k', '1')} "
                    f"--fil_h {flags.get('y', '1')} --fil_w {flags.get('x', '1')} "
                    f"--dilation_h {flags.get('l', '1')} --dilation_w {flags.get('j', '1')} "
                    f"--conv_stride_h {flags.get('u', '1')} --conv_stride_w {flags.get('v', '1')} "
                    f"--padding_h {flags.get('p', '0')} --padding_w {flags.get('q', '0')} "
                    f"--groupsize {flags.get('g', '1')} "
                    f"--perf_config={perf_config}")

    cmd = f"{rocmlir_gen} {gen_args} | {rocmlir_driver} -c --debug-only=serialize-to-isa"
    try:
        result = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=120)
        return result.stderr + result.stdout
    except Exception:
        return ""


def _parse_conv_flags(argv: list) -> dict:
    """Parse conv flag pairs like -F 1 -f NCHW into a dict."""
    flags: dict = {}
    i = 0
    while i < len(argv):
        if argv[i].startswith('-') and i + 1 < len(argv):
            key = argv[i].lstrip('-')
            flags[key] = argv[i + 1]
            i += 2
        else:
            i += 1
    return flags


def extract_all_isa_widths(build_dir: str,
                           operation: str,
                           tuning_entries: Dict[str, TuningEntry],
                           progress_label: str = "") -> Dict[str, int]:
    """Extract global_store width for each config in tuning DB."""
    results: Dict[str, int] = {}
    total = len(tuning_entries)
    for idx, (tv, entry) in enumerate(tuning_entries.items(), 1):
        if progress_label:
            print(f"  [{progress_label}] ISA extraction {idx}/{total}: {tv[:60]}...",
                  end='\r',
                  flush=True)
        isa = get_isa_for_config(build_dir, operation, tv, entry.perf_config, entry.arch,
                                 entry.num_cu, entry.num_chiplets)
        results[tv] = extract_global_store_width_from_isa(isa)
    if progress_label:
        print()
    return results


# ---------------------------------------------------------------------------
# Data merging and diff computation
# ---------------------------------------------------------------------------


@dataclass
class ComparisonRow:
    problem_config: str
    perf_config_base: str
    perf_config_feature: str
    tflops_base: float
    tflops_feature: float
    store_width_base: int
    store_width_feature: int
    diff_pct: float


def merge_results(
    base_tuning: Dict[str, TuningEntry],
    feature_tuning: Dict[str, TuningEntry],
    base_perf: Dict[str, PerfEntry],
    feature_perf: Dict[str, PerfEntry],
    base_isa: Dict[str, int],
    feature_isa: Dict[str, int],
) -> List[ComparisonRow]:
    """Merge base and feature data into comparison rows."""
    all_keys = sorted(set(base_tuning.keys()) | set(feature_tuning.keys()))
    rows: List[ComparisonRow] = []
    for tv in all_keys:
        pc_base = base_tuning[tv].perf_config if tv in base_tuning else ""
        pc_feat = feature_tuning[tv].perf_config if tv in feature_tuning else ""

        tflops_base = base_perf[tv].tflops if tv in base_perf else float('nan')
        tflops_feat = feature_perf[tv].tflops if tv in feature_perf else float('nan')

        sw_base = base_isa.get(tv, 0)
        sw_feat = feature_isa.get(tv, 0)

        diff = compute_diff_pct(tflops_base, tflops_feat)

        rows.append(
            ComparisonRow(
                problem_config=tv,
                perf_config_base=pc_base,
                perf_config_feature=pc_feat,
                tflops_base=tflops_base,
                tflops_feature=tflops_feat,
                store_width_base=sw_base,
                store_width_feature=sw_feat,
                diff_pct=diff,
            ))
    return rows


def compute_diff_pct(base_tflops: float, feature_tflops: float) -> float:
    """Compute percentage diff: ((feature - base) / base) * 100."""
    if not np.isfinite(base_tflops) or not np.isfinite(feature_tflops):
        return float('nan')
    if base_tflops == 0:
        return float('nan')
    return ((feature_tflops - base_tflops) / base_tflops) * 100.0


def sort_comparison_rows(rows: List[ComparisonRow]) -> List[ComparisonRow]:
    """Sort by Diff% descending: positive first, then negative, NaN last."""

    def sort_key(row: ComparisonRow) -> Tuple[int, float]:
        if np.isnan(row.diff_pct):
            return (2, 0.0)
        return (0, -row.diff_pct)

    return sorted(rows, key=sort_key)


# ---------------------------------------------------------------------------
# Excel report generation
# ---------------------------------------------------------------------------


@dataclass
class ExcelMetadata:
    base_branch: str = ''
    feature_branch: str = ''
    base_commit: str = ''
    feature_commit: str = ''
    hostname: str = ''
    gpu_arch: str = ''


def generate_excel(gemm_rows: List[ComparisonRow],
                   conv_rows: List[ComparisonRow],
                   output_path: str,
                   base_branch: str,
                   feature_branch: str,
                   metadata: Optional[ExcelMetadata] = None) -> None:
    """Generate an Excel workbook with Summary, GEMM, and Conv sheets."""
    if not HAS_OPENPYXL:
        print("ERROR: openpyxl is required. Install with: pip install openpyxl")
        sys.exit(1)

    if metadata is None:
        metadata = ExcelMetadata(base_branch=base_branch, feature_branch=feature_branch)

    headers = [
        'Problem Config',
        f'Winning PerfConfig ({base_branch})',
        f'Winning PerfConfig ({feature_branch})',
        f'TFlops ({base_branch})',
        f'TFlops ({feature_branch})',
        f'GlobalStore Width ({base_branch})',
        f'GlobalStore Width ({feature_branch})',
        'Diff (%)',
    ]

    df_gemm = _rows_to_dataframe(gemm_rows, headers)
    df_conv = _rows_to_dataframe(conv_rows, headers)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        _write_summary_sheet(writer, metadata, gemm_rows, conv_rows)
        df_gemm.to_excel(writer, sheet_name='GEMM', index=False)
        df_conv.to_excel(writer, sheet_name='Conv', index=False)

    _format_workbook(output_path, headers, base_branch, feature_branch)
    print(f"Excel report written to: {output_path}")


def _write_summary_sheet(writer: pd.ExcelWriter, metadata: ExcelMetadata,
                         gemm_rows: List[ComparisonRow], conv_rows: List[ComparisonRow]) -> None:
    """Write a Summary sheet with metadata and high-level stats."""
    from datetime import datetime, timezone

    gemm_finite = [r.diff_pct for r in gemm_rows if np.isfinite(r.diff_pct)]
    conv_finite = [r.diff_pct for r in conv_rows if np.isfinite(r.diff_pct)]

    summary_data = [
        ['Hostname', metadata.hostname or '(not provided)'],
        ['GPU Arch', metadata.gpu_arch or '(not provided)'],
        ['Date', datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')],
        ['', ''],
        ['Base Branch', metadata.base_branch],
        ['Base Commit', metadata.base_commit or '(not provided)'],
        ['Feature Branch', metadata.feature_branch],
        ['Feature Commit', metadata.feature_commit or '(not provided)'],
        ['', ''],
        ['GEMM Configs', len(gemm_rows)],
        ['Conv Configs', len(conv_rows)],
        ['', ''],
        ['GEMM Avg Diff (%)', f"{np.mean(gemm_finite):.2f}" if gemm_finite else 'N/A'],
        ['GEMM Improved (>0%)', sum(1 for d in gemm_finite if d > 0)],
        ['GEMM Regressed (<0%)', sum(1 for d in gemm_finite if d < 0)],
        ['GEMM Unchanged (0%)', sum(1 for d in gemm_finite if d == 0)],
        ['', ''],
        ['Conv Avg Diff (%)', f"{np.mean(conv_finite):.2f}" if conv_finite else 'N/A'],
        ['Conv Improved (>0%)', sum(1 for d in conv_finite if d > 0)],
        ['Conv Regressed (<0%)', sum(1 for d in conv_finite if d < 0)],
        ['Conv Unchanged (0%)', sum(1 for d in conv_finite if d == 0)],
    ]

    df = pd.DataFrame(summary_data, columns=['Property', 'Value'])
    df.to_excel(writer, sheet_name='Summary', index=False)


def _rows_to_dataframe(rows: List[ComparisonRow], headers: List[str]) -> pd.DataFrame:
    data = []
    for r in rows:
        data.append({
            headers[0]: r.problem_config,
            headers[1]: r.perf_config_base,
            headers[2]: r.perf_config_feature,
            headers[3]: r.tflops_base,
            headers[4]: r.tflops_feature,
            headers[5]: r.store_width_base,
            headers[6]: r.store_width_feature,
            headers[7]: r.diff_pct,
        })
    return pd.DataFrame(data, columns=headers)


def _format_workbook(path: str, headers: List[str], base_branch: str, feature_branch: str) -> None:
    """Apply formatting to the Excel workbook."""
    wb = load_workbook(path)
    bold_font = Font(bold=True)
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    right_align = Alignment(horizontal='right')

    tflops_cols = [4, 5]  # 1-indexed: TFlops base, TFlops feature
    diff_col = 8
    store_cols = [6, 7]
    numeric_cols = tflops_cols + store_cols + [diff_col]

    for ws in wb.worksheets:
        if ws.title == 'Summary':
            # Format Summary sheet: bold Property column, auto-width
            for row_idx in range(1, ws.max_row + 1):
                ws.cell(row=row_idx, column=1).font = bold_font
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 50
            continue

        for cell in ws[1]:
            cell.font = bold_font

        for col_idx in numeric_cols:
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = right_align

        for col_idx in tflops_cols:
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.number_format = '0.0000'

        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=diff_col)
            cell.number_format = '0.00"%"'
            try:
                val = cell.value
                if val is not None and np.isfinite(float(val)):
                    if float(val) > 0:
                        cell.fill = green_fill
                    elif float(val) < 0:
                        cell.fill = red_fill
            except (TypeError, ValueError):
                pass

        for col_idx in range(1, len(headers) + 1):
            max_len = len(str(ws.cell(row=1, column=col_idx).value))
            for row_idx in range(2, min(ws.max_row + 1, 20)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, min(len(str(val)), 60))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

    wb.save(path)


# ---------------------------------------------------------------------------
# Spot-check validation
# ---------------------------------------------------------------------------


def validate_excel(output_path: str,
                   gemm_rows: List[ComparisonRow],
                   conv_rows: List[ComparisonRow],
                   num_checks: int = 5) -> bool:
    """Randomly sample rows and verify Excel content matches in-memory data."""
    if not HAS_OPENPYXL:
        print("Skipping validation: openpyxl not available")
        return True

    wb = load_workbook(output_path)
    all_ok = True

    for sheet_name, rows in [('GEMM', gemm_rows), ('Conv', conv_rows)]:
        if sheet_name not in wb.sheetnames or not rows:
            continue
        ws = wb[sheet_name]
        n = min(num_checks, len(rows))
        indices = random.sample(range(len(rows)), n)
        for idx in indices:
            excel_row = idx + 2  # 1-indexed header + data
            row = rows[idx]

            excel_config = ws.cell(row=excel_row, column=1).value
            if excel_config != row.problem_config:
                print(f"  MISMATCH [{sheet_name}] row {idx}: "
                      f"config '{excel_config}' != '{row.problem_config}'")
                all_ok = False
                continue

            excel_tflops_base = ws.cell(row=excel_row, column=4).value
            excel_tflops_feat = ws.cell(row=excel_row, column=5).value
            excel_diff = ws.cell(row=excel_row, column=8).value

            if np.isfinite(row.tflops_base):
                if excel_tflops_base is None or abs(float(excel_tflops_base) -
                                                    row.tflops_base) > 0.001:
                    print(f"  MISMATCH [{sheet_name}] row {idx}: "
                          f"tflops_base {excel_tflops_base} != {row.tflops_base}")
                    all_ok = False

            if np.isfinite(row.tflops_feature):
                if excel_tflops_feat is None or abs(float(excel_tflops_feat) -
                                                    row.tflops_feature) > 0.001:
                    print(f"  MISMATCH [{sheet_name}] row {idx}: "
                          f"tflops_feature {excel_tflops_feat} != {row.tflops_feature}")
                    all_ok = False

            if np.isfinite(row.diff_pct):
                if excel_diff is None or abs(float(excel_diff) - row.diff_pct) > 0.01:
                    print(f"  MISMATCH [{sheet_name}] row {idx}: "
                          f"diff {excel_diff} != {row.diff_pct}")
                    all_ok = False

        # Check for missing TFlops and zero GlobalStore widths across ALL rows
        missing_base_tflops = 0
        missing_feat_tflops = 0
        zero_base_store = 0
        zero_feat_store = 0
        for row in rows:
            if row.perf_config_base and not np.isfinite(row.tflops_base):
                missing_base_tflops += 1
            if row.perf_config_feature and not np.isfinite(row.tflops_feature):
                missing_feat_tflops += 1
            if row.perf_config_base and row.store_width_base == 0:
                zero_base_store += 1
            if row.perf_config_feature and row.store_width_feature == 0:
                zero_feat_store += 1

        if missing_base_tflops:
            print(f"  WARNING [{sheet_name}]: {missing_base_tflops} base configs "
                  f"have perfConfig but missing TFlops")
            all_ok = False
        if missing_feat_tflops:
            print(f"  WARNING [{sheet_name}]: {missing_feat_tflops} feature configs "
                  f"have perfConfig but missing TFlops")
            all_ok = False
        if zero_base_store:
            print(f"  WARNING [{sheet_name}]: {zero_base_store} base configs "
                  f"have perfConfig but zero GlobalStore width")
        if zero_feat_store:
            print(f"  WARNING [{sheet_name}]: {zero_feat_store} feature configs "
                  f"have perfConfig but zero GlobalStore width")

    wb.close()
    if all_ok:
        print("Validation passed: random spot-checks all match.")
    return all_ok


def validate_excel_against_sources(
    excel_path: str,
    base_gemm_tsv: str,
    base_gemm_csv: str,
    feat_gemm_tsv: str,
    feat_gemm_csv: str,
    base_conv_tsv: str,
    base_conv_csv: str,
    feat_conv_tsv: str,
    feat_conv_csv: str,
    base_branch: str,
    feature_branch: str,
    num_checks: int = 10,
) -> bool:
    """Independent validation: re-read source TSV/CSV files and verify
    the Excel contains matching data. This catches bugs where in-memory
    data might be consistent but diverges from the source files."""
    if not HAS_OPENPYXL:
        print("Skipping source validation: openpyxl not available")
        return True

    print(f"Independent validation: sampling {num_checks} rows per sheet "
          f"from source files...")

    wb = load_workbook(excel_path)
    all_ok = True

    for sheet_name, operation, b_tsv, b_csv, f_tsv, f_csv in [
        ('GEMM', 'gemm', base_gemm_tsv, base_gemm_csv, feat_gemm_tsv, feat_gemm_csv),
        ('Conv', 'conv', base_conv_tsv, base_conv_csv, feat_conv_tsv, feat_conv_csv),
    ]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue

        base_tuning = parse_tuning_tsv(b_tsv)
        feat_tuning = parse_tuning_tsv(f_tsv)
        base_perf = parse_perf_csv(b_csv, operation)
        feat_perf = parse_perf_csv(f_csv, operation)

        data_rows = ws.max_row - 1
        n = min(num_checks, data_rows)
        indices = random.sample(range(2, ws.max_row + 1), n)

        for excel_row in indices:
            config = ws.cell(row=excel_row, column=1).value
            if config is None:
                continue

            excel_pc_base = ws.cell(row=excel_row, column=2).value or ""
            excel_pc_feat = ws.cell(row=excel_row, column=3).value or ""
            excel_tflops_base = ws.cell(row=excel_row, column=4).value
            excel_tflops_feat = ws.cell(row=excel_row, column=5).value
            excel_diff = ws.cell(row=excel_row, column=8).value

            # Check perfConfig against tuning DB
            if config in base_tuning:
                src_pc = base_tuning[config].perf_config
                if str(excel_pc_base) != str(src_pc):
                    print(f"  MISMATCH [{sheet_name}] row {excel_row}: "
                          f"base perfConfig '{excel_pc_base}' != source '{src_pc}'")
                    all_ok = False

            if config in feat_tuning:
                src_pc = feat_tuning[config].perf_config
                if str(excel_pc_feat) != str(src_pc):
                    print(f"  MISMATCH [{sheet_name}] row {excel_row}: "
                          f"feature perfConfig '{excel_pc_feat}' != source '{src_pc}'")
                    all_ok = False

            # Check TFlops against perf CSV
            if config in base_perf and np.isfinite(base_perf[config].tflops):
                src_tflops = base_perf[config].tflops
                if excel_tflops_base is not None:
                    if abs(float(excel_tflops_base) - src_tflops) > 0.001:
                        print(f"  MISMATCH [{sheet_name}] row {excel_row}: "
                              f"base TFlops {excel_tflops_base} != source {src_tflops}")
                        all_ok = False

            if config in feat_perf and np.isfinite(feat_perf[config].tflops):
                src_tflops = feat_perf[config].tflops
                if excel_tflops_feat is not None:
                    if abs(float(excel_tflops_feat) - src_tflops) > 0.001:
                        print(f"  MISMATCH [{sheet_name}] row {excel_row}: "
                              f"feature TFlops {excel_tflops_feat} != source {src_tflops}")
                        all_ok = False

            # Verify diff calculation independently
            if (excel_tflops_base is not None and excel_tflops_feat is not None and
                    excel_diff is not None):
                try:
                    b = float(excel_tflops_base)
                    f = float(excel_tflops_feat)
                    d = float(excel_diff)
                    if np.isfinite(b) and np.isfinite(f) and b != 0:
                        expected_diff = ((f - b) / b) * 100.0
                        if abs(d - expected_diff) > 0.01:
                            print(f"  MISMATCH [{sheet_name}] row {excel_row}: "
                                  f"diff {d} != expected {expected_diff:.2f} "
                                  f"(base={b}, feat={f})")
                            all_ok = False
                except (TypeError, ValueError):
                    pass

        # Check for missing TFlops: configs in tuning DB should have perf data
        missing_base = 0
        missing_feat = 0
        zero_store_base = 0
        zero_store_feat = 0
        for row_idx in range(2, ws.max_row + 1):
            config = ws.cell(row=row_idx, column=1).value
            if config is None:
                continue
            pc_base = ws.cell(row=row_idx, column=2).value or ""
            pc_feat = ws.cell(row=row_idx, column=3).value or ""
            tfl_base = ws.cell(row=row_idx, column=4).value
            tfl_feat = ws.cell(row=row_idx, column=5).value
            sw_base = ws.cell(row=row_idx, column=6).value
            sw_feat = ws.cell(row=row_idx, column=7).value

            if pc_base and tfl_base is None:
                missing_base += 1
            if pc_feat and tfl_feat is None:
                missing_feat += 1
            if pc_base and (sw_base is None or sw_base == 0):
                zero_store_base += 1
            if pc_feat and (sw_feat is None or sw_feat == 0):
                zero_store_feat += 1

        if missing_base:
            print(f"  WARNING [{sheet_name}]: {missing_base}/{ws.max_row - 1} base configs "
                  f"have perfConfig but missing TFlops")
            all_ok = False
        if missing_feat:
            print(f"  WARNING [{sheet_name}]: {missing_feat}/{ws.max_row - 1} feature configs "
                  f"have perfConfig but missing TFlops")
            all_ok = False
        if zero_store_base:
            print(f"  WARNING [{sheet_name}]: {zero_store_base}/{ws.max_row - 1} base configs "
                  f"have perfConfig but zero GlobalStore width")
        if zero_store_feat:
            print(f"  WARNING [{sheet_name}]: {zero_store_feat}/{ws.max_row - 1} feature configs "
                  f"have perfConfig but zero GlobalStore width")

        # Verify sorting: Diff column should be descending (positive first, NaN last)
        diffs = []
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=8).value
            diffs.append(val)

        finite_diffs = [float(d) for d in diffs if d is not None and np.isfinite(float(d))]
        if len(finite_diffs) > 1:
            for i in range(len(finite_diffs) - 1):
                if finite_diffs[i] < finite_diffs[i + 1]:
                    print(f"  SORT ERROR [{sheet_name}]: "
                          f"row {i+2} diff {finite_diffs[i]} < "
                          f"row {i+3} diff {finite_diffs[i+1]} "
                          f"(should be descending)")
                    all_ok = False
                    break

    wb.close()
    if all_ok:
        print("Independent validation passed: Excel matches source files "
              "and sorting is correct.")
    else:
        print("WARNING: Independent validation found mismatches!")
    return all_ok


# ---------------------------------------------------------------------------
# CLI and main
# ---------------------------------------------------------------------------


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="A/B performance comparison between two rocMLIR branches")

    parser.add_argument('--feature-branch',
                        default='swapOperands2',
                        help='Feature branch name (default: swapOperands2)')
    parser.add_argument('--base-branch',
                        default='develop',
                        help='Base branch name (default: develop)')

    parser.add_argument('--feature-build-dir',
                        required=True,
                        help='Build directory for the feature branch')
    parser.add_argument('--base-build-dir',
                        required=True,
                        help='Build directory for the base branch')

    parser.add_argument('--feature-gemm-tsv',
                        required=True,
                        help='Tuning TSV for feature branch GEMM')
    parser.add_argument('--feature-conv-tsv',
                        required=True,
                        help='Tuning TSV for feature branch Conv')
    parser.add_argument('--base-gemm-tsv', required=True, help='Tuning TSV for base branch GEMM')
    parser.add_argument('--base-conv-tsv', required=True, help='Tuning TSV for base branch Conv')

    parser.add_argument('--feature-gemm-csv',
                        required=True,
                        help='Perf CSV for feature branch GEMM')
    parser.add_argument('--feature-conv-csv',
                        required=True,
                        help='Perf CSV for feature branch Conv')
    parser.add_argument('--base-gemm-csv', required=True, help='Perf CSV for base branch GEMM')
    parser.add_argument('--base-conv-csv', required=True, help='Perf CSV for base branch Conv')

    parser.add_argument('--gemm-configs', help='Path to tier1-gemm-configs file (optional)')
    parser.add_argument('--conv-configs', help='Path to tier1-conv-configs file (optional)')

    parser.add_argument('--output',
                        '-o',
                        default='perf_comparison.xlsx',
                        help='Output Excel file path')
    parser.add_argument('--skip-isa',
                        action='store_true',
                        help='Skip ISA extraction (set all widths to 0)')
    parser.add_argument('--validate',
                        action='store_true',
                        default=True,
                        help='Run spot-check validation after generating Excel')
    parser.add_argument('--no-validate', action='store_true', help='Skip spot-check validation')

    parser.add_argument('--base-commit', default='', help='Git commit hash for the base branch')
    parser.add_argument('--feature-commit',
                        default='',
                        help='Git commit hash for the feature branch')
    parser.add_argument('--hostname',
                        default='',
                        help='Hostname of the machine running the comparison')
    parser.add_argument('--gpu-arch', default='', help='GPU architecture (e.g. gfx950)')

    return parser


def main(args=None) -> int:
    parser = build_parser()
    parsed = parser.parse_args(args)

    print(f"=== Performance Comparison: {parsed.base_branch} vs {parsed.feature_branch} ===\n")

    # Parse tuning DBs
    print("Parsing tuning databases...")
    base_gemm_tuning = parse_tuning_tsv(parsed.base_gemm_tsv)
    feat_gemm_tuning = parse_tuning_tsv(parsed.feature_gemm_tsv)
    base_conv_tuning = parse_tuning_tsv(parsed.base_conv_tsv)
    feat_conv_tuning = parse_tuning_tsv(parsed.feature_conv_tsv)
    print(f"  Base GEMM: {len(base_gemm_tuning)} configs, "
          f"Feature GEMM: {len(feat_gemm_tuning)} configs")
    print(f"  Base Conv: {len(base_conv_tuning)} configs, "
          f"Feature Conv: {len(feat_conv_tuning)} configs")

    # Parse perf CSVs
    print("Parsing performance CSVs...")
    base_gemm_perf = parse_perf_csv(parsed.base_gemm_csv, 'gemm')
    feat_gemm_perf = parse_perf_csv(parsed.feature_gemm_csv, 'gemm')
    base_conv_perf = parse_perf_csv(parsed.base_conv_csv, 'conv')
    feat_conv_perf = parse_perf_csv(parsed.feature_conv_csv, 'conv')
    print(f"  Base GEMM perf: {len(base_gemm_perf)} entries, "
          f"Feature GEMM perf: {len(feat_gemm_perf)} entries")
    print(f"  Base Conv perf: {len(base_conv_perf)} entries, "
          f"Feature Conv perf: {len(feat_conv_perf)} entries")

    # ISA extraction
    if parsed.skip_isa:
        print("Skipping ISA extraction (--skip-isa)")
        base_gemm_isa: Dict[str, int] = {}
        feat_gemm_isa: Dict[str, int] = {}
        base_conv_isa: Dict[str, int] = {}
        feat_conv_isa: Dict[str, int] = {}
    else:
        print("Extracting ISA global_store widths...")
        base_gemm_isa = extract_all_isa_widths(parsed.base_build_dir, 'gemm', base_gemm_tuning,
                                               f"{parsed.base_branch}/gemm")
        feat_gemm_isa = extract_all_isa_widths(parsed.feature_build_dir, 'gemm', feat_gemm_tuning,
                                               f"{parsed.feature_branch}/gemm")
        base_conv_isa = extract_all_isa_widths(parsed.base_build_dir, 'conv', base_conv_tuning,
                                               f"{parsed.base_branch}/conv")
        feat_conv_isa = extract_all_isa_widths(parsed.feature_build_dir, 'conv', feat_conv_tuning,
                                               f"{parsed.feature_branch}/conv")

    # Merge and compute diffs
    print("Merging results and computing diffs...")
    gemm_rows = sort_comparison_rows(
        merge_results(base_gemm_tuning, feat_gemm_tuning, base_gemm_perf, feat_gemm_perf,
                      base_gemm_isa, feat_gemm_isa))
    conv_rows = sort_comparison_rows(
        merge_results(base_conv_tuning, feat_conv_tuning, base_conv_perf, feat_conv_perf,
                      base_conv_isa, feat_conv_isa))

    # Generate Excel
    print(f"Generating Excel report: {parsed.output}")
    metadata = ExcelMetadata(
        base_branch=parsed.base_branch,
        feature_branch=parsed.feature_branch,
        base_commit=parsed.base_commit,
        feature_commit=parsed.feature_commit,
        hostname=parsed.hostname,
        gpu_arch=parsed.gpu_arch,
    )
    generate_excel(gemm_rows,
                   conv_rows,
                   parsed.output,
                   parsed.base_branch,
                   parsed.feature_branch,
                   metadata=metadata)

    # Validate
    if not parsed.no_validate:
        print("Running spot-check validation...")
        validate_excel(parsed.output, gemm_rows, conv_rows)

    print("\nDone!")
    return 0


if __name__ == '__main__':
    sys.exit(main())
