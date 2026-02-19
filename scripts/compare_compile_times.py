#!/usr/bin/env python3
"""
Compare compile times from with vs without sched_group runs.
Reads compile_time_without_sched_group.txt and compile_time_with_sched_group.txt,
produces comparison with compile times, diff %, grid_size, and block_size; writes to Excel.
Grid/block are obtained by running rocmlir-gen + rocmlir-driver once per perf_config (cached).
"""

import argparse
import json
import math
import re
import subprocess
import sys
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import load_workbook
except ImportError:
    raise SystemExit("pip install pandas openpyxl")

COMPILE_MS_RE = re.compile(r"compile_ms=([\d.eE+-]+)")
# From "func.func @rock_gemm(...) attributes { block_size = 64 : i32, ... grid_size = 4096 : i32 ...}"
GRID_SIZE_RE = re.compile(r"grid_size\s*=\s*(\d+)\s*:\s*i32")
BLOCK_SIZE_RE = re.compile(r"block_size\s*=\s*(\d+)\s*:\s*i32")


def parse_compile_time_file(path: Path) -> dict[str, float]:
    """Parse a compile time file: PerfConfig -> compile_ms (seconds or ms as float)."""
    result = {}
    with open(path) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = line.split("\t", 1)
            perf_config = parts[0].strip()
            if len(parts) < 2 or parts[1].strip() == "N/A":
                continue
            m = COMPILE_MS_RE.search(parts[1])
            if m:
                result[perf_config] = float(m.group(1))
    return result


def get_grid_block_for_perf_config(
    perf_config: str,
    rocmlir_gen: Path,
    rocmlir_driver: Path,
    arch: str,
    m: int,
    n: int,
    k: int,
    g: int,
) -> tuple[int | None, int | None]:
    """Run rocmlir-gen | rocmlir-driver and parse grid_size, block_size from IR. Returns (grid_size, block_size) or (None, None) on failure."""
    gen_cmd = [
        str(rocmlir_gen),
        "-operation",
        "gemm",
        "-t",
        "f16",
        "-out_datatype",
        "f16",
        "--arch",
        arch,
        "--num_cu",
        "256",
        "--num_chiplets",
        "8",
        "-g",
        str(g),
        "-m",
        str(m),
        "-k",
        str(k),
        "-n",
        str(n),
        "-transA=False",
        "-transB=False",
        f"--perf_config={perf_config}",
    ]
    driver_cmd = [
        str(rocmlir_driver),
        "--kernel-pipeline=gpu",
        "--mlir-print-ir-after=rock-gemm-to-gridwise",
    ]
    try:
        gen_proc = subprocess.Popen(
            gen_cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.DEVNULL,
            text=True,
        )
        driver_proc = subprocess.run(
            driver_cmd,
            stdin=gen_proc.stdout,
            capture_output=True,
            text=True,
            timeout=120,
        )
        if gen_proc.wait() != 0:
            return (None, None)
        ir_output = driver_proc.stdout + driver_proc.stderr
        grid_m = GRID_SIZE_RE.search(ir_output)
        block_m = BLOCK_SIZE_RE.search(ir_output)
        grid_size = int(grid_m.group(1)) if grid_m else None
        block_size = int(block_m.group(1)) if block_m else None
        return (grid_size, block_size)
    except (subprocess.TimeoutExpired, FileNotFoundError, ValueError):
        return (None, None)


def load_grid_block_cache(cache_path: Path) -> dict[str, dict[str, int | None]]:
    """Load cache: perf_config -> {grid_size: int|None, block_size: int|None}."""
    if not cache_path.exists():
        return {}
    try:
        with open(cache_path) as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError):
        return {}


def save_grid_block_cache(cache_path: Path, cache: dict[str, dict[str, int | None]]) -> None:
    """Save cache."""
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    with open(cache_path, "w") as f:
        json.dump(cache, f, indent=0)


def main() -> None:
    parser = argparse.ArgumentParser(description="Compare compile times with/without sched_group")
    parser.add_argument(
        "--without",
        type=Path,
        default=Path(__file__).resolve().parent.parent / "build" /
        "compile_time_without_sched_group.txt",
        help="Path to compile_time_without_sched_group.txt",
    )
    parser.add_argument(
        "--with-sched",
        type=Path,
        default=Path(__file__).resolve().parent.parent / "build" /
        "compile_time_with_sched_group.txt",
        help="Path to compile_time_with_sched_group.txt",
    )
    parser.add_argument(
        "--with-iglp",
        type=Path,
        default=Path(__file__).resolve().parent.parent / "build" / "compile_time_with_iglp.txt",
        help="Path to compile_time_with_iglp.txt (optional)",
    )
    parser.add_argument(
        "--no-iglp",
        action="store_true",
        help="Do not load or include IGLP columns",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=Path(__file__).resolve().parent.parent / "build" / "compile_time_comparison.xlsx",
        help="Output Excel file path",
    )
    parser.add_argument(
        "--all-configs",
        action="store_true",
        help=
        "Include rows for every PerfConfig (use N/A where missing); default: only configs with both values",
    )
    parser.add_argument(
        "--build-dir",
        type=Path,
        default=Path(__file__).resolve().parent.parent / "build",
        help="Build directory containing bin/rocmlir-gen and bin/rocmlir-driver",
    )
    parser.add_argument(
        "--grid-block-cache",
        type=Path,
        default=None,
        help=
        "JSON file to cache grid_size/block_size per perf_config (default: build/perf_config_grid_block.json)",
    )
    parser.add_argument(
        "--no-cache",
        action="store_true",
        help="Ignore and overwrite grid/block cache; recompute all",
    )
    parser.add_argument(
        "--skip-grid-block",
        action="store_true",
        help=
        "Do not run rocmlir-gen/driver; omit grid_size and block_size columns (or use cache only)",
    )
    parser.add_argument(
        "--arch",
        default="gfx950:sramecc+:xnack-",
        help="Architecture for rocmlir-gen",
    )
    parser.add_argument(
        "-m",
        "--gemm-m",
        type=int,
        default=256,
        help="GEMM M dimension",
    )
    parser.add_argument(
        "-n",
        "--gemm-n",
        type=int,
        default=4096,
        help="GEMM N dimension",
    )
    parser.add_argument(
        "-k",
        "--gemm-k",
        type=int,
        default=14336,
        help="GEMM K dimension",
    )
    parser.add_argument(
        "-g",
        "--gemm-g",
        type=int,
        default=1,
        help="GEMM G (batch) dimension",
    )
    args = parser.parse_args()
    if args.grid_block_cache is None:
        args.grid_block_cache = args.build_dir / "perf_config_grid_block.json"

    without = parse_compile_time_file(args.without)
    with_sched = parse_compile_time_file(args.with_sched)
    with_iglp: dict[str, float] = {}
    if not args.no_iglp and args.with_iglp.exists():
        with_iglp = parse_compile_time_file(args.with_iglp)

    if args.all_configs:
        all_configs = sorted(set(without) | set(with_sched))
    else:
        all_configs = sorted(set(without) & set(with_sched))

    rocmlir_gen = args.build_dir / "bin" / "rocmlir-gen"
    rocmlir_driver = args.build_dir / "bin" / "rocmlir-driver"
    grid_block_cache = {} if args.no_cache else load_grid_block_cache(args.grid_block_cache)
    cache_updated = False

    rows = []
    for i, perf_config in enumerate(all_configs):
        wo = without.get(perf_config)
        wi = with_sched.get(perf_config)

        # Store as numbers for Excel; use NaN where missing
        wo_val = float(wo) if wo is not None else math.nan
        wi_val = float(wi) if wi is not None else math.nan
        if wo is not None and wi is not None and wo != 0:
            # Base = without_sched_group; store as percentage number (e.g. 5.23 for +5.23%)
            pct_val = ((wi - wo) / wo) * 100
        else:
            pct_val = math.nan

        iglp_val = float(with_iglp[perf_config]) if perf_config in with_iglp else math.nan
        if wo is not None and wo != 0 and perf_config in with_iglp:
            pct_iglp_val = ((with_iglp[perf_config] - wo) / wo) * 100
        else:
            pct_iglp_val = math.nan

        grid_size: int | float = math.nan
        block_size: int | float = math.nan
        if not args.skip_grid_block:
            cached = grid_block_cache.get(perf_config)
            if cached is not None and not args.no_cache:
                grid_size = cached["grid_size"] if cached.get("grid_size") is not None else math.nan
                block_size = cached["block_size"] if cached.get(
                    "block_size") is not None else math.nan
            else:
                gs, bs = get_grid_block_for_perf_config(
                    perf_config,
                    rocmlir_gen,
                    rocmlir_driver,
                    args.arch,
                    args.gemm_m,
                    args.gemm_n,
                    args.gemm_k,
                    args.gemm_g,
                )
                grid_size = gs if gs is not None else math.nan
                block_size = bs if bs is not None else math.nan
                grid_block_cache[perf_config] = {"grid_size": gs, "block_size": bs}
                cache_updated = True
            if (i + 1) % 50 == 0:
                print(f"Grid/block: {i + 1}/{len(all_configs)} configs processed...",
                      file=sys.stderr)

        row_dict: dict = {
            "PerfConfig": perf_config,
            "CompileTime_without_schedGroup_ms": wo_val,
            "CompileTime_with_schedGroup_ms": wi_val,
            "CompileTime_difference_pct": pct_val,
        }
        if with_iglp:
            row_dict["CompileTime_with_IGLP_ms"] = iglp_val
            row_dict["CompileTime_difference_IGLP_pct"] = pct_iglp_val
        row_dict["grid_size"] = grid_size
        row_dict["block_size"] = block_size
        rows.append(row_dict)

    if cache_updated:
        save_grid_block_cache(args.grid_block_cache, grid_block_cache)
        print(f"Updated grid/block cache: {args.grid_block_cache}", file=sys.stderr)

    df = pd.DataFrame(rows)
    # Sort by difference column: largest to smallest (NaN last)
    df = df.sort_values("CompileTime_difference_pct", ascending=False, na_position="last")
    args.output.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(args.output, index=False, sheet_name="Compare", engine="openpyxl")

    # Apply number formats so Excel reads and displays values correctly
    wb = load_workbook(args.output)
    ws = wb["Compare"]
    # A=PerfConfig, B=without, C=with_sched, D=diff_pct, [E=with_IGLP, F=diff_IGLP_pct], grid_size, block_size
    for row in range(2, len(rows) + 2):
        c = 2
        ws.cell(row=row, column=c).number_format = "0.00"  # CompileTime without
        c += 1
        ws.cell(row=row, column=c).number_format = "0.00"  # CompileTime with_sched
        c += 1
        ws.cell(row=row, column=c).number_format = '0.00"%"'  # diff sched %
        c += 1
        if with_iglp:
            ws.cell(row=row, column=c).number_format = "0.00"  # CompileTime with_IGLP
            c += 1
            ws.cell(row=row, column=c).number_format = '0.00"%"'  # diff IGLP %
            c += 1
        ws.cell(row=row, column=c).number_format = "0"  # grid_size
        c += 1
        ws.cell(row=row, column=c).number_format = "0"  # block_size
    wb.save(args.output)
    wb.close()

    print(f"Wrote {len(rows)} rows to {args.output}")


if __name__ == "__main__":
    main()
