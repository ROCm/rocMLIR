#!/usr/bin/env python3
"""
Run tier1-gemm benchmarks on schedGroup branch using one or more tuning DBs,
then add TFlops and % diff vs Develop to the existing comparison Excel.
Applies conditional formatting (red < 0, green > 0) to all % diff columns including
Pct_Diff_vs_Develop (schedGroup vs develop).

Usage (from repo root):
  python3 scripts/add_iglp_to_comparison.py [--no-run] [--add-tuning TSV ...] [--add-csv CSV ...]
  --no-run       Skip benchmarks; only merge existing CSVs into Excel.
  --add-tuning   Additional tuning DB(s) to run and add (e.g. iglp_with_one_wave.tsv).
  --add-csv      Existing perf result CSV(s) to merge (no benchmark run). TFlops column name = filename stem + _TFlops.
"""

import argparse
import os
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_BUILD_DIR = REPO_ROOT / "build"
TIER1_GEMM_CONFIGS = REPO_ROOT / "mlir/utils/performance/configs/tier1-gemm-configs"
DEFAULT_EXCEL = DEFAULT_BUILD_DIR / "tflops_comparison_develop_vs_schedGroup.xlsx"

# (tsv_basename, output_csv_basename, TFlops column name in Excel)
TUNING_DBS = [
    ("iglp.tsv", "iglp_tier1_gemm_results.csv", "iglpOpt_TFlops"),
    ("iglp_with_one_wave.tsv", "iglp_with_one_wave_tier1_gemm_results.csv",
     "iglp_with_one_wave_TFlops"),
    ("iglp_with_one_wave_200.tsv", "iglp_with_one_wave_200_tier1_gemm_results.csv",
     "iglp_with_one_wave_200_TFlops"),
]

GEMM_MERGE_KEY = [
    "DataType",
    "OutDataType",
    "Chip",
    "numCU",
    "numChiplets",
    "TransA",
    "TransB",
    "G",
    "M",
    "K",
    "N",
    "ScaledGemm",
    "ScaleADtype",
    "ScaleBDtype",
    "TransScaleA",
    "TransScaleB",
]

RED_FILL = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
GREEN_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")


def run_cmd(cmd, cwd=None, check=True):
    import subprocess
    cwd = cwd or Path.cwd()
    print(f"  $ {' '.join(str(x) for x in cmd)}")
    r = subprocess.run(cmd, cwd=str(cwd))
    if check and r.returncode != 0:
        raise RuntimeError(f"Command failed with exit code {r.returncode}")
    return r.returncode


def get_current_branch(repo_root):
    import subprocess
    return subprocess.check_output(
        ["git", "rev-parse", "--abbrev-ref", "HEAD"],
        cwd=repo_root,
        text=True,
    ).strip()


def run_perf_benchmark(build_dir, tsv_path, configs_path, output_csv):
    """Run perfRunner with given tuning TSV from build_dir (schedGroup branch)."""
    build_dir = Path(build_dir)
    configs_rel = os.path.relpath(configs_path, build_dir)
    cmd = [
        sys.executable,
        str(build_dir / "bin" / "perfRunner.py"),
        "-t",
        str(tsv_path),
        "-c",
        str(configs_rel),
        "--batch_mlir",
        "--op",
        "gemm",
        "-o",
        str(output_csv),
    ]
    run_cmd(cmd, cwd=build_dir)


def pct_diff_col(tflops_col):
    """Pct diff column name for a TFlops column (vs Develop)."""
    base = tflops_col.replace("_TFlops", "")
    return f"Pct_Diff_{base}_vs_Develop"


def add_tuning_runs_and_format(excel_path, build_dir, tuning_dbs, run_benchmark, extra_csvs=None):
    """extra_csvs: list of (csv_path, tflops_col_name) for existing CSVs to merge (no run)."""
    build_dir = Path(build_dir)
    excel_path = Path(excel_path)
    extra_csvs = extra_csvs or []

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    df = pd.read_excel(excel_path, sheet_name=0)
    for col in GEMM_MERGE_KEY + ["Develop_TFlops"]:
        if col not in df.columns:
            raise ValueError(f"Excel missing column {col}")

    # Only run/merge tuning DBs that are not already in the Excel
    tuning_dbs = [(t, c, n) for (t, c, n) in tuning_dbs if n not in df.columns]
    # Only add extra CSVs that are not already in the Excel
    extra_csvs = [(p, n) for (p, n) in extra_csvs if n not in df.columns]

    # Ensure we're on schedGroup if we'll run any benchmark
    if run_benchmark and tuning_dbs and any((build_dir / tsv).exists() for tsv, _, _ in tuning_dbs):
        branch = get_current_branch(REPO_ROOT)
        if branch != "schedGroup":
            print(f"Checking out schedGroup (current: {branch})")
            run_cmd(["git", "checkout", "schedGroup"], cwd=REPO_ROOT)

    def merge_one_csv(df, csv_path, tflops_col):
        csv_path = Path(csv_path)
        if not csv_path.exists():
            print(f"  Skip {tflops_col}: CSV not found {csv_path}")
            return df
        df_run = pd.read_csv(csv_path)
        for col in GEMM_MERGE_KEY + ["TFlops"]:
            if col not in df_run.columns:
                raise ValueError(f"{csv_path} missing column {col}")
        df_run = df_run[GEMM_MERGE_KEY + ["TFlops"]].rename(columns={"TFlops": tflops_col})
        df = df.merge(df_run, on=GEMM_MERGE_KEY, how="left")
        pct_col = pct_diff_col(tflops_col)
        ref = df["Develop_TFlops"]
        other = df[tflops_col]
        df[pct_col] = np.where(
            pd.notna(ref) & (ref != 0),
            (other - ref) / ref * 100.0,
            np.nan,
        )
        return df

    for tsv_basename, csv_basename, tflops_col in tuning_dbs:
        tsv_path = build_dir / tsv_basename
        csv_path = build_dir / csv_basename

        if run_benchmark and tsv_path.exists():
            print(f"Running benchmark with {tsv_basename} -> {csv_basename}")
            run_perf_benchmark(build_dir, tsv_path, TIER1_GEMM_CONFIGS, csv_path)

        df = merge_one_csv(df, csv_path, tflops_col)

    for csv_path, tflops_col in extra_csvs:
        print(f"Merging existing CSV: {csv_path} -> {tflops_col}")
        df = merge_one_csv(df, Path(csv_path), tflops_col)

    # Column order: key cols, then Develop_TFlops, then pairs of (other TFlops, Pct_Diff_*)
    key_cols = [c for c in GEMM_MERGE_KEY if c in df.columns]
    develop_cols = ["Develop_TFlops", "schedGroup_TFlops", "Pct_Diff_vs_Develop"]
    rest = [c for c in df.columns if c not in key_cols and c not in develop_cols]
    # Group rest into (TFlops, Pct_Diff) pairs and keep order
    ordered = key_cols + [c for c in develop_cols if c in df.columns]
    seen = set(ordered)
    for c in rest:
        if c in seen:
            continue
        seen.add(c)
        ordered.append(c)
        if c.endswith("_TFlops"):
            pct = pct_diff_col(c)
            if pct in df.columns and pct not in seen:
                seen.add(pct)
                ordered.append(pct)
    ordered += [c for c in rest if c not in seen]
    df = df[[c for c in ordered if c in df.columns]]

    df.to_excel(excel_path, index=False, sheet_name="Sheet1")

    # Apply red/green to every Pct_Diff* column
    wb = load_workbook(excel_path)
    ws = wb.active
    pct_columns = [c for c in df.columns if c.startswith("Pct_Diff")]
    max_row = len(df) + 1

    for col_name in pct_columns:
        col_idx = list(df.columns).index(col_name) + 1
        col_letter = get_column_letter(col_idx)
        cell_range = f"{col_letter}2:{col_letter}{max_row}"
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL),
        )
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="greaterThan", formula=["0"], fill=GREEN_FILL),
        )

    wb.save(excel_path)
    print(f"Updated: {excel_path}")
    print(f"  TFlops columns: Develop_TFlops, schedGroup_TFlops" +
          "".join(f", {c}" for c in df.columns
                  if c.endswith("_TFlops") and c not in ("Develop_TFlops", "schedGroup_TFlops")))
    print(f"  Formatted % diff columns (red < 0, green > 0): {', '.join(pct_columns)}")


def main():
    parser = argparse.ArgumentParser(description="Add tuning run results to comparison Excel.")
    parser.add_argument("--no-run",
                        action="store_true",
                        help="Skip benchmarks; only merge existing CSVs.")
    parser.add_argument("--excel",
                        type=Path,
                        default=DEFAULT_EXCEL,
                        help="Path to comparison Excel file.")
    parser.add_argument("--build-dir",
                        type=Path,
                        default=DEFAULT_BUILD_DIR,
                        help="Build directory.")
    parser.add_argument(
        "--add-tuning",
        type=str,
        nargs="*",
        default=None,
        metavar="TSV",
        help=
        "Extra tuning DB(s) to run and add (e.g. iglp_with_one_wave.tsv). Default: all in TUNING_DBS list.",
    )
    parser.add_argument(
        "--add-csv",
        type=Path,
        nargs="*",
        default=None,
        metavar="CSV",
        help=
        "Existing perf result CSV(s) to merge (no benchmark run). Column name = filename stem + _TFlops.",
    )
    args = parser.parse_args()

    if args.add_tuning is not None:
        tuning_dbs = []
        for tsv in args.add_tuning:
            tsv = tsv if tsv.endswith(".tsv") else f"{tsv}.tsv"
            base = tsv.replace(".tsv", "")
            csv = f"{base}_tier1_gemm_results.csv"
            tflops_col = f"{base}_TFlops" if base != "iglp" else "iglpOpt_TFlops"
            tuning_dbs.append((tsv, csv, tflops_col))
    else:
        tuning_dbs = [t for t in TUNING_DBS if (args.build_dir / t[0]).exists()]

    extra_csvs = []
    if args.add_csv:
        for p in args.add_csv:
            p = Path(p)
            stem = p.stem
            tflops_col = f"{stem}_TFlops"
            extra_csvs.append((p.resolve(), tflops_col))

    if not tuning_dbs and not extra_csvs:
        print("Nothing to add (no tuning DBs or --add-csv, or already in Excel).")
        add_tuning_runs_and_format(args.excel,
                                   args.build_dir, [],
                                   run_benchmark=False,
                                   extra_csvs=[])
        return

    add_tuning_runs_and_format(
        args.excel,
        args.build_dir,
        tuning_dbs,
        run_benchmark=not args.no_run,
        extra_csvs=extra_csvs,
    )


if __name__ == "__main__":
    main()
